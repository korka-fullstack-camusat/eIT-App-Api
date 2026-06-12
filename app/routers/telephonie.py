from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional
from datetime import date
import csv, io
from ..database import get_db
from ..models.telephonie import NumeroSIM, SiteGSM, Vehicule, AffectationSIM, CategorieSimEnum, StatutSimEnum, ImportGlobalLog
from ..models.facture import FactureTelecom, LigneFacture
from sqlalchemy import and_
from ..schemas.telephonie import (
    NumeroSIMCreate, NumeroSIMUpdate, NumeroSIMOut,
    SiteGSMCreate, SiteGSMOut,
    VehiculeCreate, VehiculeOut,
    AffectationSIMCreate, AffectationSIMOut, DesaffectationSIMIn,
    ImportGlobalLogOut,
)
from ..models.user import User
from ..services.auth_service import require_editor

router = APIRouter(prefix="/api/telephonie", tags=["Téléphonie"])

# ── SIMs ──────────────────────────────────────────────────────────────────────

@router.get("/sims", response_model=list[NumeroSIMOut])
def list_sims(
    categorie: Optional[CategorieSimEnum] = None,
    statut:    Optional[StatutSimEnum] = None,
    search:    Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(NumeroSIM).options(joinedload(NumeroSIM.affectation_active))
    if categorie: q = q.filter(NumeroSIM.categorie == categorie)
    if statut:    q = q.filter(NumeroSIM.statut == statut)
    if search:    q = q.filter(NumeroSIM.numero.ilike(f"%{search}%"))
    sims = q.order_by(NumeroSIM.numero).all()

    type_ligne = "MOBILE" if categorie == CategorieSimEnum.EMPLOYE else None
    fact_map = _latest_facture_map(db, type_ligne=type_ligne)
    result = []
    for s in sims:
        item = NumeroSIMOut.model_validate(s)
        item.derniere_facture = fact_map.get(s.id)
        result.append(item)
    return result


@router.post("/sims", response_model=NumeroSIMOut, status_code=201)
def create_sim(data: NumeroSIMCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    if db.query(NumeroSIM).filter(NumeroSIM.numero == data.numero).first():
        raise HTTPException(400, "Ce numéro existe déjà")
    obj = NumeroSIM(**data.model_dump())
    db.add(obj); db.commit(); db.refresh(obj)
    return obj


@router.get("/sims/export")
def export_sims(
    categorie: Optional[CategorieSimEnum] = None,
    statut:    Optional[StatutSimEnum]    = None,
    search:    Optional[str]              = None,
    db: Session = Depends(get_db),
):
    q = db.query(NumeroSIM).options(joinedload(NumeroSIM.affectation_active))
    if categorie: q = q.filter(NumeroSIM.categorie == categorie)
    if statut:    q = q.filter(NumeroSIM.statut == statut)
    if search:    q = q.filter(NumeroSIM.numero.ilike(f"%{search}%"))
    sims = q.order_by(NumeroSIM.numero).all()

    rows = []
    rows.append(["Numero", "Categorie", "Operateur", "Statut", "Description",
                 "Affecte a", "Matricule", "Date affectation"])
    for s in sims:
        aff = s.affectation_active
        rows.append([
            s.numero,
            s.categorie.value,
            s.operateur or "",
            s.statut.value,
            s.description or "",
            aff.employee_nom        if aff and aff.employee_nom        else "",
            aff.employee_matricule  if aff and aff.employee_matricule  else "",
            aff.date_debut.strftime("%d/%m/%Y") if aff else "",
        ])

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerows(rows)

    # BOM UTF-8 (chr(0xFEFF)) pour ouverture correcte dans Excel
    content = chr(0xFEFF) + output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=numeros_sim.csv"},
    )


@router.get("/sims/export-excel")
def export_sims_excel(
    categorie: Optional[CategorieSimEnum] = None,
    statut:    Optional[StatutSimEnum]    = None,
    search:    Optional[str]              = None,
    cols:      Optional[str]              = None,
    db: Session = Depends(get_db),
):
    """Export Excel stylisé des numéros SIM."""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(NumeroSIM).options(joinedload(NumeroSIM.affectation_active))
    if categorie: q = q.filter(NumeroSIM.categorie == categorie)
    if statut:    q = q.filter(NumeroSIM.statut == statut)
    if search:    q = q.filter(NumeroSIM.numero.ilike(f"%{search}%"))
    sims = q.order_by(NumeroSIM.numero).all()

    CAT_LABELS  = {"EMPLOYE": "Employé", "M2M_SITE": "M2M Site", "M2M_VEHICULE": "M2M Véhicule"}
    STAT_LABELS = {"ACTIVE": "Active", "INACTIVE": "Inactive", "SUSPENDUE": "Suspendue",
                   "RESILIE": "Résilié", "CEDE": "Cédé"}

    ALL_COLS = [
        ("numero",    "Numéro",          lambda s, a: s.numero or ""),
        ("imsi",      "IMSI",            lambda s, a: s.imsi or ""),
        ("sim_matricule", "Matricule",   lambda s, a: s.matricule or ""),
        ("beneficiaire",  "Bénéficiaire",lambda s, a: s.beneficiaire or ""),
        ("service",       "Service",     lambda s, a: s.service or ""),
        ("business_line", "BL",          lambda s, a: s.business_line or ""),
        ("fonction",      "Fonction",    lambda s, a: s.fonction or ""),
        ("categorie", "Catégorie",       lambda s, a: CAT_LABELS.get(s.categorie, s.categorie or "")),
        ("operateur", "Opérateur",       lambda s, a: s.operateur or ""),
        ("statut",    "Statut",          lambda s, a: STAT_LABELS.get(s.statut, s.statut or "")),
        ("affecte",   "Affecté à",       lambda s, a: (a.employee_nom or (f"Site #{a.site_id}" if a.site_id else f"Véhicule #{a.vehicule_id}")) if a else ""),
        ("matricule", "Matricule",       lambda s, a: a.employee_matricule or "" if a else ""),
        ("date_aff",  "Date affectation",lambda s, a: a.date_debut.strftime("%d/%m/%Y") if a else ""),
        ("desc",      "Description",     lambda s, a: s.description or ""),
    ]
    selected_keys = set(cols.split(",")) if cols else {c[0] for c in ALL_COLS}
    columns = [(k, lbl, fn) for k, lbl, fn in ALL_COLS if k in selected_keys]

    BLUE_HDR = "1B3D6F"; WHITE = "FFFFFF"; ROW_EVEN = "EEF4FF"; BORDER_COL = "C5D3E8"
    thin = Side(style="thin", color=BORDER_COL)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook(); ws = wb.active; ws.title = "Numéros SIM"

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    tc = ws.cell(row=1, column=1, value="INVENTAIRE DES NUMÉROS SIM")
    tc.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=BLUE_HDR)
    tc.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 32

    # Sous-titre
    parts = [f"Exporté le {datetime.now().strftime('%d/%m/%Y')}", f"{len(sims)} numéro(s)"]
    if statut:    parts.append(f"Statut : {STAT_LABELS.get(statut, statut)}")
    if categorie: parts.append(f"Catégorie : {CAT_LABELS.get(categorie, categorie)}")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sc = ws.cell(row=2, column=1, value="  |  ".join(parts))
    sc.font = Font(name="Calibri", italic=True, size=9, color="5B7DB1")
    sc.fill = PatternFill("solid", fgColor="D9E6F7")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    HDR_ROW = 4
    for ci, (_, lbl, _fn) in enumerate(columns, start=1):
        c = ws.cell(row=HDR_ROW, column=ci, value=lbl)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border_all
    ws.row_dimensions[HDR_ROW].height = 24; ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    for ri, s in enumerate(sims):
        row_num = HDR_ROW + 1 + ri
        aff = s.affectation_active
        row_fill = PatternFill("solid", fgColor=(ROW_EVEN if ri % 2 == 0 else WHITE))
        for ci, (_, _lbl, fn) in enumerate(columns, start=1):
            c = ws.cell(row=row_num, column=ci, value=fn(s, aff))
            c.fill = row_fill; c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", horizontal="left", indent=1); c.border = border_all
        ws.row_dimensions[row_num].height = 18

    for ci, (_, lbl, fn) in enumerate(columns, start=1):
        max_len = max((len(str(fn(s, s.affectation_active))) for s in sims), default=0)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(lbl) + 2, min(max_len + 2, 40))

    ws.oddFooter.center.text = "&\"Calibri\"&8 CAMUSAT — Numéros SIM  |  Page &P / &N"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=numeros_sim.xlsx"},
    )


@router.post("/sims/import")
async def import_sims(file: UploadFile = File(...), db: Session = Depends(get_db), _: User = Depends(require_editor)):
    CAT_MAP = {
        "EMPLOYE": "EMPLOYE", "EMPLOYÉ": "EMPLOYE", "EMPLOYE": "EMPLOYE",
        "M2M_SITE": "M2M_SITE", "M2M SITE": "M2M_SITE", "SITE": "M2M_SITE",
        "M2M_VEHICULE": "M2M_VEHICULE", "M2M VEHICULE": "M2M_VEHICULE",
        "M2M VÉHICULE": "M2M_VEHICULE", "VEHICULE": "M2M_VEHICULE",
    }

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    def norm(s: str) -> str:
        import unicodedata
        s = unicodedata.normalize("NFD", s)
        s = "".join(c for c in s if unicodedata.category(c) != "Mn")
        return s.strip().upper().replace(" ", "_")

    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    created, updated, errors = 0, 0, []

    for i, raw_row in enumerate(reader, start=2):
        row = {norm(k): (v.strip() if v else "") for k, v in raw_row.items()}

        numero = row.get("NUMERO", row.get("NUM", "")).strip()
        if not numero:
            errors.append({"ligne": i, "message": "Numéro manquant"}); continue

        cat_raw = row.get("CATEGORIE", "EMPLOYE").strip().upper()
        cat_val = CAT_MAP.get(norm(cat_raw))
        if not cat_val:
            errors.append({"ligne": i, "message": f"Catégorie inconnue : '{cat_raw}'"}); continue

        operateur   = row.get("OPERATEUR", row.get("OPERATEUR_", "")) or None
        description = row.get("DESCRIPTION", "") or None

        existing = db.query(NumeroSIM).filter(NumeroSIM.numero == numero).first()
        if existing:
            existing.operateur   = operateur   if operateur   else existing.operateur
            existing.description = description if description else existing.description
            updated += 1
        else:
            db.add(NumeroSIM(
                numero=numero,
                categorie=CategorieSimEnum(cat_val),
                operateur=operateur,
                description=description,
            ))
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "errors": errors,
            "total_lignes": created + updated + len(errors)}


@router.patch("/sims/{sim_id}", response_model=NumeroSIMOut)
def update_sim(sim_id: int, data: NumeroSIMUpdate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(NumeroSIM).filter(NumeroSIM.id == sim_id).first()
    if not obj: raise HTTPException(404, "SIM introuvable")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(obj, k, v)
    db.commit(); db.refresh(obj)
    return obj


@router.delete("/sims/{sim_id}", status_code=204)
def delete_sim(sim_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(NumeroSIM).filter(NumeroSIM.id == sim_id).first()
    if not obj: raise HTTPException(404, "SIM introuvable")
    db.delete(obj); db.commit()


# ── Affectations SIM ──────────────────────────────────────────────────────────

@router.post("/sims/{sim_id}/affecter", response_model=AffectationSIMOut, status_code=201)
def affecter_sim(sim_id: int, data: AffectationSIMCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    active = db.query(AffectationSIM).filter(
        AffectationSIM.sim_id == sim_id, AffectationSIM.is_active == True
    ).first()
    if active:
        who = active.employee_nom or (f"Site #{active.site_id}" if active.site_id else f"Véhicule #{active.vehicule_id}")
        raise HTTPException(409, f"Ce numéro est déjà affecté à {who}. Veuillez d'abord le désaffecter.")

    payload = data.model_dump()
    payload["sim_id"] = sim_id
    obj = AffectationSIM(**payload)
    db.add(obj); db.commit(); db.refresh(obj)
    return obj


@router.patch("/sims/{sim_id}/desaffecter", response_model=AffectationSIMOut)
def desaffecter_sim(sim_id: int, data: DesaffectationSIMIn, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    active = db.query(AffectationSIM).filter(
        AffectationSIM.sim_id == sim_id, AffectationSIM.is_active == True
    ).first()
    if not active:
        raise HTTPException(404, "Aucune affectation active pour ce numéro")
    active.is_active = False
    active.date_fin  = date.today()
    active.motif_fin = data.motif
    db.commit(); db.refresh(active)
    return active


@router.get("/sims/{sim_id}/historique", response_model=list[AffectationSIMOut])
def historique_sim(sim_id: int, db: Session = Depends(get_db)):
    return (
        db.query(AffectationSIM)
        .filter(AffectationSIM.sim_id == sim_id)
        .order_by(AffectationSIM.date_debut.desc())
        .all()
    )


# ── Sites GSM ─────────────────────────────────────────────────────────────────

def _site_sim_map(db: Session) -> dict:
    """Retourne un dict {site_id: (sim_numero, sim_id, operateur)} pour toutes les affectations actives."""
    rows = (
        db.query(AffectationSIM.site_id, NumeroSIM.numero, NumeroSIM.id, NumeroSIM.operateur)
        .join(NumeroSIM, NumeroSIM.id == AffectationSIM.sim_id)
        .filter(
            AffectationSIM.is_active == True,
            AffectationSIM.site_id.isnot(None),
        )
        .all()
    )
    return {r.site_id: (r.numero, r.id, r.operateur) for r in rows}


def _latest_facture_map(db: Session, type_ligne: Optional[str] = None) -> dict:
    """Retourne un dict {sim_id: {mois, annee, operateur, montant_ttc}} correspondant
    à la facture la plus récente (mois/année) reconnue pour chaque numéro SIM.

    Si `type_ligne` est fourni, ne considère que les lignes de facture dont le
    `type_ligne` correspond (ex: "MOBILE" pour les numéros employés)."""
    from sqlalchemy import func as _func

    base = db.query(LigneFacture).filter(LigneFacture.sim_id.isnot(None))
    if type_ligne:
        base = base.filter(LigneFacture.type_ligne == type_ligne)
    base_ids = base.with_entities(LigneFacture.id).subquery()

    latest_sub = (
        db.query(
            LigneFacture.sim_id.label("sim_id"),
            _func.max(FactureTelecom.annee * 100 + FactureTelecom.mois).label("ym"),
        )
        .join(FactureTelecom, LigneFacture.facture_id == FactureTelecom.id)
        .filter(LigneFacture.id.in_(base_ids))
        .group_by(LigneFacture.sim_id)
        .subquery()
    )

    q = (
        db.query(LigneFacture, FactureTelecom)
        .join(FactureTelecom, LigneFacture.facture_id == FactureTelecom.id)
        .join(
            latest_sub,
            and_(
                LigneFacture.sim_id == latest_sub.c.sim_id,
                (FactureTelecom.annee * 100 + FactureTelecom.mois) == latest_sub.c.ym,
            ),
        )
    )
    if type_ligne:
        q = q.filter(LigneFacture.type_ligne == type_ligne)
    rows = q.all()
    return {
        l.sim_id: {
            "mois":        f.mois,
            "annee":       f.annee,
            "operateur":   f.operateur,
            "montant_ttc": float(l.montant_ttc) if l.montant_ttc is not None else (float(l.montant) if l.montant is not None else None),
        }
        for l, f in rows
    }


@router.get("/sites", response_model=list[SiteGSMOut])
def list_sites(db: Session = Depends(get_db)):
    sites    = db.query(SiteGSM).order_by(SiteGSM.nom).all()
    sim_map  = _site_sim_map(db)
    fact_map = _latest_facture_map(db)
    result   = []
    for s in sites:
        item = SiteGSMOut.model_validate(s)
        sim_info = sim_map.get(s.id)
        if sim_info:
            item.sim_numero = sim_info[0]
            item.sim_operateur = sim_info[2]
            item.derniere_facture = fact_map.get(sim_info[1])
        result.append(item)
    return result


def _sims_total_for_period(db: Session, mois: int, annee: int, categorie: Optional[CategorieSimEnum] = None) -> dict:
    """Coût total facturé pour les numéros SIM (toutes catégories, ou une
    catégorie donnée), pour un mois/année donné(e)."""
    from sqlalchemy import func as _func

    montant_expr = _func.coalesce(LigneFacture.montant_ttc, LigneFacture.montant)
    q = (
        db.query(
            _func.coalesce(_func.sum(montant_expr), 0),
            _func.count(LigneFacture.id),
        )
        .join(FactureTelecom, LigneFacture.facture_id == FactureTelecom.id)
        .join(NumeroSIM, NumeroSIM.id == LigneFacture.sim_id)
        .filter(FactureTelecom.mois == mois, FactureTelecom.annee == annee)
    )
    if categorie:
        q = q.filter(NumeroSIM.categorie == categorie)
    total, nb = q.first()
    return {"mois": mois, "annee": annee, "total": float(total or 0), "nombre_numeros": int(nb or 0)}


@router.get("/sims/stats/periodes")
def sims_stats_periodes(categorie: Optional[CategorieSimEnum] = None, db: Session = Depends(get_db)):
    """Liste des mois/années pour lesquels une facture contient au moins un
    numéro SIM (toutes catégories, ou une catégorie donnée)."""
    q = (
        db.query(FactureTelecom.annee, FactureTelecom.mois)
        .join(LigneFacture, LigneFacture.facture_id == FactureTelecom.id)
        .join(NumeroSIM, NumeroSIM.id == LigneFacture.sim_id)
    )
    if categorie:
        q = q.filter(NumeroSIM.categorie == categorie)
    rows = q.distinct().order_by(FactureTelecom.annee.desc(), FactureTelecom.mois.desc()).all()
    return [{"mois": r.mois, "annee": r.annee} for r in rows]


@router.get("/sims/stats/mensuel")
def sims_stats_mensuel(mois: int, annee: int, categorie: Optional[CategorieSimEnum] = None, db: Session = Depends(get_db)):
    """Coût total des numéros SIM pour un mois/année donné(e)."""
    return _sims_total_for_period(db, mois, annee, categorie)


@router.get("/sims/stats/ecart")
def sims_stats_ecart(
    mois1: int, annee1: int, mois2: int, annee2: int,
    categorie: Optional[CategorieSimEnum] = None, db: Session = Depends(get_db),
):
    """Écart de coût total des numéros SIM entre deux mois/années."""
    p1 = _sims_total_for_period(db, mois1, annee1, categorie)
    p2 = _sims_total_for_period(db, mois2, annee2, categorie)
    ecart = p2["total"] - p1["total"]
    ecart_pct = (ecart / p1["total"] * 100) if p1["total"] else None
    return {"periode1": p1, "periode2": p2, "ecart": ecart, "ecart_pct": ecart_pct}


@router.get("/sims/stats/evolution")
def sims_stats_evolution(categorie: Optional[CategorieSimEnum] = None, db: Session = Depends(get_db)):
    """Évolution du coût total des numéros SIM mois par mois, avec l'écart
    (valeur et %) par rapport au mois précédent, pour toutes les périodes
    disponibles."""
    q = (
        db.query(FactureTelecom.mois, FactureTelecom.annee)
        .join(LigneFacture, LigneFacture.facture_id == FactureTelecom.id)
        .join(NumeroSIM, NumeroSIM.id == LigneFacture.sim_id)
    )
    if categorie:
        q = q.filter(NumeroSIM.categorie == categorie)
    rows = q.distinct().all()
    periodes = sorted(set((r.annee, r.mois) for r in rows))

    result = []
    prev_total = None
    for annee, mois in periodes:
        p = _sims_total_for_period(db, mois, annee, categorie)
        if prev_total is not None:
            ecart = p["total"] - prev_total
            ecart_pct = (ecart / prev_total * 100) if prev_total else None
        else:
            ecart = None
            ecart_pct = None
        result.append({**p, "ecart": ecart, "ecart_pct": ecart_pct})
        prev_total = p["total"]
    return result


def _sites_total_for_period(db: Session, mois: int, annee: int) -> dict:
    """Coût total facturé pour les numéros SIM des sites RMS (catégorie
    M2M_SITE), pour un mois/année donné(e)."""
    from sqlalchemy import func as _func

    montant_expr = _func.coalesce(LigneFacture.montant_ttc, LigneFacture.montant)
    row = (
        db.query(
            _func.coalesce(_func.sum(montant_expr), 0),
            _func.count(LigneFacture.id),
        )
        .join(FactureTelecom, LigneFacture.facture_id == FactureTelecom.id)
        .join(NumeroSIM, NumeroSIM.id == LigneFacture.sim_id)
        .filter(
            FactureTelecom.mois  == mois,
            FactureTelecom.annee == annee,
            NumeroSIM.categorie  == CategorieSimEnum.M2M_SITE,
        )
        .first()
    )
    total, nb = row
    return {"mois": mois, "annee": annee, "total": float(total or 0), "nombre_numeros": int(nb or 0)}


@router.get("/sites/stats/periodes")
def sites_stats_periodes(db: Session = Depends(get_db)):
    """Liste des mois/années pour lesquels une facture contient au moins un
    numéro SIM de site RMS (catégorie M2M_SITE) — pour alimenter les sélecteurs
    de période des statistiques."""
    rows = (
        db.query(FactureTelecom.annee, FactureTelecom.mois)
        .join(LigneFacture, LigneFacture.facture_id == FactureTelecom.id)
        .join(NumeroSIM, NumeroSIM.id == LigneFacture.sim_id)
        .filter(NumeroSIM.categorie == CategorieSimEnum.M2M_SITE)
        .distinct()
        .order_by(FactureTelecom.annee.desc(), FactureTelecom.mois.desc())
        .all()
    )
    return [{"mois": r.mois, "annee": r.annee} for r in rows]


@router.get("/sites/stats/mensuel")
def sites_stats_mensuel(mois: int, annee: int, db: Session = Depends(get_db)):
    """Coût total des sites RMS pour un mois/année donné(e)."""
    return _sites_total_for_period(db, mois, annee)


@router.get("/sites/stats/ecart")
def sites_stats_ecart(
    mois1: int, annee1: int, mois2: int, annee2: int, db: Session = Depends(get_db),
):
    """Écart de coût total des sites RMS entre deux mois/années."""
    p1 = _sites_total_for_period(db, mois1, annee1)
    p2 = _sites_total_for_period(db, mois2, annee2)
    ecart = p2["total"] - p1["total"]
    ecart_pct = (ecart / p1["total"] * 100) if p1["total"] else None
    return {"periode1": p1, "periode2": p2, "ecart": ecart, "ecart_pct": ecart_pct}


@router.get("/sites/stats/evolution")
def sites_stats_evolution(db: Session = Depends(get_db)):
    """Évolution du coût total des sites RMS mois par mois, avec l'écart
    (valeur et %) par rapport au mois précédent, pour toutes les périodes
    disponibles."""
    rows = (
        db.query(FactureTelecom.mois, FactureTelecom.annee)
        .join(LigneFacture, LigneFacture.facture_id == FactureTelecom.id)
        .join(NumeroSIM, NumeroSIM.id == LigneFacture.sim_id)
        .filter(NumeroSIM.categorie == CategorieSimEnum.M2M_SITE)
        .distinct()
        .all()
    )
    periodes = sorted(set((r.annee, r.mois) for r in rows))

    result = []
    prev_total = None
    for annee, mois in periodes:
        p = _sites_total_for_period(db, mois, annee)
        if prev_total is not None:
            ecart = p["total"] - prev_total
            ecart_pct = (ecart / prev_total * 100) if prev_total else None
        else:
            ecart = None
            ecart_pct = None
        result.append({**p, "ecart": ecart, "ecart_pct": ecart_pct})
        prev_total = p["total"]
    return result


@router.get("/sites/{site_id}/facturation")
def get_site_facturation(site_id: int, db: Session = Depends(get_db)):
    """Historique mensuel de facturation du SIM actuellement affecté à ce site,
    reconstitué automatiquement à partir des factures télécom importées."""
    site = db.query(SiteGSM).filter(SiteGSM.id == site_id).first()
    if not site:
        raise HTTPException(404, "Site introuvable")

    aff = (
        db.query(AffectationSIM)
        .filter(AffectationSIM.site_id == site_id, AffectationSIM.is_active == True)
        .first()
    )
    if not aff:
        return {"sim_numero": None, "lignes": []}

    sim = db.query(NumeroSIM).filter(NumeroSIM.id == aff.sim_id).first()

    rows = (
        db.query(LigneFacture, FactureTelecom)
        .join(FactureTelecom, LigneFacture.facture_id == FactureTelecom.id)
        .filter(LigneFacture.sim_id == aff.sim_id)
        .order_by(FactureTelecom.annee.desc(), FactureTelecom.mois.desc())
        .all()
    )

    return {
        "sim_numero": sim.numero if sim else None,
        "lignes": [
            {
                "mois":      f.mois,
                "annee":     f.annee,
                "operateur": f.operateur,
                "montant":     float(l.montant)     if l.montant     is not None else None,
                "montant_ttc": float(l.montant_ttc) if l.montant_ttc is not None else None,
            }
            for l, f in rows
        ],
    }


@router.post("/sites", response_model=SiteGSMOut, status_code=201)
def create_site(data: SiteGSMCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = SiteGSM(**data.model_dump())
    db.add(obj); db.commit(); db.refresh(obj)
    return obj


@router.get("/sites/export")
def export_sites(db: Session = Depends(get_db)):
    sites   = db.query(SiteGSM).order_by(SiteGSM.nom).all()
    sim_map = _site_sim_map(db)
    rows = [["SiteID", "IMSI", "Nom du site", "Localisation", "Numero SIM"]]
    for s in sites:
        rows.append([
            s.code_site    or "",
            s.imsi         or "",
            s.nom,
            s.localisation or "",
            (sim_map.get(s.id) or (None,))[0] or "",
        ])
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerows(rows)
    content = chr(0xFEFF) + output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=sites_gsm.csv"},
    )


@router.get("/sites/export-excel")
def export_sites_excel(
    filter_sim:  Optional[str]  = None,   # "affecte" | "non_affecte"
    search:      Optional[str]  = None,
    cols:        Optional[str]  = None,
    date_debut:  Optional[date] = None,   # filtre sur created_at
    date_fin:    Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Export Excel stylisé des sites GSM."""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    sites   = db.query(SiteGSM).order_by(SiteGSM.nom).all()
    sim_map = _site_sim_map(db)

    # Enrichir avec sim_numero
    class SiteRow:
        def __init__(self, s):
            self.id          = s.id
            self.code_site   = s.code_site
            self.imsi        = s.imsi
            self.nom         = s.nom
            self.localisation = s.localisation
            self.created_at  = s.created_at
            self.sim_numero  = (sim_map.get(s.id) or (None,))[0]

    rows_data = [SiteRow(s) for s in sites]

    if filter_sim == "affecte":
        rows_data = [r for r in rows_data if r.sim_numero]
    elif filter_sim == "non_affecte":
        rows_data = [r for r in rows_data if not r.sim_numero]
    if date_debut:
        rows_data = [r for r in rows_data if r.created_at and r.created_at.date() >= date_debut]
    if date_fin:
        rows_data = [r for r in rows_data if r.created_at and r.created_at.date() <= date_fin]
    if search:
        q = search.lower()
        rows_data = [r for r in rows_data if
            q in (r.nom or "").lower() or q in (r.code_site or "").lower() or
            q in (r.imsi or "").lower() or q in (r.sim_numero or "").lower() or
            q in (r.localisation or "").lower()]

    ALL_COLS = [
        ("code_site",    "SiteID",          lambda r: r.code_site or ""),
        ("imsi",         "IMSI",            lambda r: r.imsi or ""),
        ("nom",          "Nom du site",     lambda r: r.nom or ""),
        ("localisation", "Localisation",    lambda r: r.localisation or ""),
        ("sim_numero",   "Numéro SIM",      lambda r: r.sim_numero or ""),
        ("statut_sim",   "Statut SIM",      lambda r: "Affecté" if r.sim_numero else "Non affecté"),
        ("created_at",   "Date création",   lambda r: r.created_at.strftime("%d/%m/%Y") if r.created_at else ""),
    ]
    selected_keys = set(cols.split(",")) if cols else {c[0] for c in ALL_COLS}
    columns = [(k, lbl, fn) for k, lbl, fn in ALL_COLS if k in selected_keys]

    BLUE_HDR = "1B3D6F"; WHITE = "FFFFFF"; ROW_EVEN = "EEF4FF"; BORDER_COL = "C5D3E8"
    thin = Side(style="thin", color=BORDER_COL)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook(); ws = wb.active; ws.title = "Sites GSM"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    tc = ws.cell(row=1, column=1, value="INVENTAIRE DES SITES RMS")
    tc.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=BLUE_HDR)
    tc.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 32

    parts = [f"Exporté le {datetime.now().strftime('%d/%m/%Y')}"]
    if date_debut: parts.append(f"Du {date_debut.strftime('%d/%m/%Y')}")
    if date_fin:   parts.append(f"au {date_fin.strftime('%d/%m/%Y')}")
    if filter_sim == "affecte":       parts.append("Avec SIM affecté")
    elif filter_sim == "non_affecte": parts.append("Sans SIM affecté")
    parts.append(f"{len(rows_data)} site(s)")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sc = ws.cell(row=2, column=1, value="  |  ".join(parts))
    sc.font = Font(name="Calibri", italic=True, size=9, color="5B7DB1")
    sc.fill = PatternFill("solid", fgColor="D9E6F7")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    HDR_ROW = 4
    for ci, (_, lbl, _fn) in enumerate(columns, start=1):
        c = ws.cell(row=HDR_ROW, column=ci, value=lbl)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border_all
    ws.row_dimensions[HDR_ROW].height = 24; ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    for ri, r in enumerate(rows_data):
        row_num = HDR_ROW + 1 + ri
        row_fill = PatternFill("solid", fgColor=(ROW_EVEN if ri % 2 == 0 else WHITE))
        for ci, (_, _lbl, fn) in enumerate(columns, start=1):
            c = ws.cell(row=row_num, column=ci, value=fn(r))
            c.fill = row_fill; c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", horizontal="left", indent=1); c.border = border_all
        ws.row_dimensions[row_num].height = 18

    for ci, (_, lbl, fn) in enumerate(columns, start=1):
        max_len = max((len(fn(r)) for r in rows_data), default=0)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(lbl) + 2, min(max_len + 2, 40))

    ws.oddFooter.center.text = "&\"Calibri\"&8 CAMUSAT — Sites RMS  |  Page &P / &N"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)

    fname = "sites_gsm"
    if date_debut: fname += f"_du_{date_debut}"
    if date_fin:   fname += f"_au_{date_fin}"
    fname += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/sites/import")
async def import_sites(file: UploadFile = File(...), db: Session = Depends(get_db), _: User = Depends(require_editor)):
    """
    Accepte deux formats de colonnes (séparateur ;) :
    - Nouveau format CAMUSAT : Numéro ; IMSI ; SITES ID ; NOMS SITES
    - Ancien format           : SiteID ; Nom du site ; Localisation ; Description

    Si la colonne "Numéro" est présente et contient un numéro SIM existant,
    une affectation est créée automatiquement entre ce SIM et le site.

    Le format .xlsx (ex : feuilles "RMS_Orange" / "RMS_Free" du fichier de
    suivi flotte) est également accepté : les feuilles contenant les colonnes
    Numéro / IMSI / SITES ID / NOMS SITES sont détectées automatiquement.
    """
    import unicodedata as _ud
    from datetime import date as _date

    content = await file.read()

    if (file.filename or "").lower().endswith(".xlsx"):
        return _import_sites_xlsx(content, db)

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    def norm(s: str) -> str:
        s = _ud.normalize("NFD", s)
        s = "".join(c for c in s if _ud.category(c) != "Mn")
        return s.strip().upper().replace(" ", "_").replace("'", "").replace("'", "")

    # Détection automatique du séparateur
    first_line = text.split("\n")[0]
    sep = ";" if ";" in first_line else "\t"

    reader = csv.DictReader(io.StringIO(text), delimiter=sep)
    created, updated, affecte, errors = 0, 0, 0, []

    for i, raw_row in enumerate(reader, start=2):
        row = {norm(k): (v.strip() if v else "") for k, v in raw_row.items()}

        # ── Colonnes NOMS SITES / NOM_DU_SITE / NOM ──────────────────────────
        nom = (
            row.get("NOMS_SITES") or row.get("NOMS_SITE") or
            row.get("NOM_SITES")  or row.get("NOM_DU_SITE") or
            row.get("NOM")        or ""
        ).strip()
        if not nom:
            errors.append({"ligne": i, "message": "Nom du site manquant"}); continue

        # ── Colonnes SITES ID / SITEID / CODE_SITE ───────────────────────────
        code_site = (
            row.get("SITES_ID") or row.get("SITEID") or
            row.get("SITE_ID")  or row.get("CODE_SITE") or ""
        ) or None

        # ── IMSI ──────────────────────────────────────────────────────────────
        imsi = row.get("IMSI", "") or None

        # ── Numéro SIM (nouveau format) ───────────────────────────────────────
        sim_numero = (
            row.get("NUMERO") or row.get("NUM") or row.get("NUMEROS") or ""
        ).strip() or None

        # ── Localisation / Description (ancien format) ─────────────────────
        localisation = row.get("LOCALISATION", "") or None
        description  = row.get("DESCRIPTION", "")  or None

        # ── Créer ou mettre à jour le site ────────────────────────────────────
        existing = (
            db.query(SiteGSM).filter(SiteGSM.code_site == code_site).first()
            if code_site else None
        ) or db.query(SiteGSM).filter(SiteGSM.nom == nom).first()

        if existing:
            if code_site:    existing.code_site    = code_site
            if imsi:         existing.imsi         = imsi
            if localisation: existing.localisation = localisation
            if description:  existing.description  = description
            site_obj = existing
            updated += 1
        else:
            site_obj = SiteGSM(
                nom=nom, code_site=code_site, imsi=imsi,
                localisation=localisation, description=description,
            )
            db.add(site_obj)
            db.flush()   # obtenir l'id sans commit
            created += 1

        # ── Auto-affectation SIM ──────────────────────────────────────────────
        if sim_numero:
            sim_obj = db.query(NumeroSIM).filter(NumeroSIM.numero == sim_numero).first()
            if sim_obj:
                # Vérifier si une affectation active existe déjà pour ce site
                existing_aff = db.query(AffectationSIM).filter(
                    AffectationSIM.site_id == site_obj.id,
                    AffectationSIM.is_active == True,
                ).first()
                if not existing_aff:
                    # Désactiver toute affectation active de ce SIM
                    old_aff = db.query(AffectationSIM).filter(
                        AffectationSIM.sim_id == sim_obj.id,
                        AffectationSIM.is_active == True,
                    ).first()
                    if old_aff:
                        old_aff.is_active = False
                        old_aff.date_fin  = _date.today()

                    db.add(AffectationSIM(
                        sim_id     = sim_obj.id,
                        site_id    = site_obj.id,
                        date_debut = _date.today(),
                        is_active  = True,
                    ))
                    affecte += 1

    db.commit()
    return {
        "created": created, "updated": updated,
        "affecte": affecte, "errors": errors,
        "total_lignes": created + updated + len(errors),
    }


def _import_sites_xlsx(content: bytes, db: Session):
    """Import depuis le fichier de suivi flotte (.xlsx) — feuilles
    'RMS_Orange' / 'RMS_Free' (ou toute feuille contenant les colonnes
    Numéro / IMSI / SITES ID / NOMS SITES). Crée/maj les sites RMS, les
    numéros SIM correspondants (catégorie M2M_SITE) et leur affectation."""
    import io, re, unicodedata
    from datetime import date as _date
    from openpyxl import load_workbook

    def norm(s) -> str:
        s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
        return re.sub(r"[^A-Z0-9]", "", s.upper())

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Fichier Excel illisible.")

    created, updated, affecte, sims_crees, errors = 0, 0, 0, 0, []

    for ws in wb.worksheets:
        sheet_norm = norm(ws.title)
        operateur = "Orange" if "ORANGE" in sheet_norm else "Free" if "FREE" in sheet_norm else None

        # Repérer la ligne d'en-tête (celle qui contient "NOMS SITES")
        header_row_idx, headers = None, {}
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=4, values_only=True), start=1):
            normalized = [norm(v) if v is not None else "" for v in row]
            if "NOMSSITES" in normalized or "NOMSITE" in normalized:
                header_row_idx = r_idx
                headers = {h: c for c, h in enumerate(normalized) if h}
                break
        if header_row_idx is None:
            continue  # feuille non concernée (ex : "Equipes pointage", "Feuil1"…)

        def col(row, *names):
            for name in names:
                idx = headers.get(name)
                if idx is not None and idx < len(row):
                    v = row[idx]
                    if isinstance(v, str):
                        v = v.strip()
                    if v not in (None, ""):
                        return v
            return None

        for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            if all(v is None for v in row):
                continue

            nom = col(row, "NOMSSITES", "NOMSITE", "NOM")
            if not nom:
                continue
            nom = str(nom).strip()

            code_site = col(row, "SITESID", "SITEID", "CODESITE")
            code_site = str(code_site).strip() if code_site is not None else None

            imsi_val = col(row, "IMSI")
            imsi = str(int(imsi_val)) if isinstance(imsi_val, float) else (str(imsi_val).strip() if imsi_val is not None else None)

            num_val = col(row, "NUMERO", "NUM")
            sim_numero = str(int(num_val)) if isinstance(num_val, float) else (str(num_val).strip() if num_val is not None else None)

            # ── Créer ou mettre à jour le site ──────────────────────────────────
            existing = (
                db.query(SiteGSM).filter(SiteGSM.code_site == code_site).first()
                if code_site else None
            ) or db.query(SiteGSM).filter(SiteGSM.nom == nom).first()

            if existing:
                if code_site: existing.code_site = code_site
                if imsi:       existing.imsi      = imsi
                site_obj = existing
                updated += 1
            else:
                site_obj = SiteGSM(nom=nom, code_site=code_site, imsi=imsi)
                db.add(site_obj)
                db.flush()
                created += 1

            # ── Créer le numéro SIM si besoin, puis l'affecter au site ──────────
            if sim_numero:
                sim_obj = db.query(NumeroSIM).filter(NumeroSIM.numero == sim_numero).first()
                if not sim_obj:
                    sim_obj = NumeroSIM(
                        numero=sim_numero, imsi=imsi,
                        categorie=CategorieSimEnum.M2M_SITE,
                        operateur=operateur,
                    )
                    db.add(sim_obj)
                    db.flush()
                    sims_crees += 1
                else:
                    if operateur and not sim_obj.operateur:
                        sim_obj.operateur = operateur
                    if imsi and not sim_obj.imsi:
                        sim_obj.imsi = imsi

                existing_aff = db.query(AffectationSIM).filter(
                    AffectationSIM.site_id == site_obj.id,
                    AffectationSIM.is_active == True,
                ).first()
                if not existing_aff:
                    old_aff = db.query(AffectationSIM).filter(
                        AffectationSIM.sim_id == sim_obj.id,
                        AffectationSIM.is_active == True,
                    ).first()
                    if old_aff:
                        old_aff.is_active = False
                        old_aff.date_fin  = _date.today()

                    db.add(AffectationSIM(
                        sim_id     = sim_obj.id,
                        site_id    = site_obj.id,
                        date_debut = _date.today(),
                        is_active  = True,
                    ))
                    affecte += 1

    relinked = _relink_lignes_facture(db)

    db.commit()
    return {
        "created": created, "updated": updated,
        "affecte": affecte, "sims_crees": sims_crees, "relinked": relinked, "errors": errors,
        "total_lignes": created + updated + len(errors),
    }


def _import_mobiles_xlsx(content: bytes, db: Session):
    """Import depuis la feuille 'ORANGE-mobiles' du fichier de suivi flotte
    (.xlsx) — crée/met à jour les numéros SIM employés (catégorie EMPLOYE)
    avec matricule / bénéficiaire / service / business line / fonction, et
    crée l'affectation employé correspondante si elle n'existe pas déjà.
    La partie facturation (colonnes mensuelles) n'est PAS traitée ici."""
    import io, re, unicodedata
    from datetime import date as _date
    from openpyxl import load_workbook

    def norm(s) -> str:
        s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
        return re.sub(r"[^A-Z0-9]", "", s.upper())

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Fichier Excel illisible.")

    created, updated, affecte, errors = 0, 0, 0, []

    if "ORANGE-mobiles" not in wb.sheetnames:
        return {"created": 0, "updated": 0, "affecte": 0, "errors": []}

    ws = wb["ORANGE-mobiles"]

    # Repérer la ligne d'en-tête (celle qui contient "MATRICULE" et "BENEFICIAIRE")
    header_row_idx = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        normalized = [norm(v) if v is not None else "" for v in row]
        if "MATRICULE" in normalized and "BENEFICIAIRE" in normalized:
            header_row_idx = r_idx
            break
    if header_row_idx is None:
        return {"created": 0, "updated": 0, "affecte": 0, "errors": ["Feuille ORANGE-mobiles : en-tête introuvable"]}

    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if all(v is None for v in row):
            continue

        matricule     = row[0] if len(row) > 0 else None
        beneficiaire  = row[1] if len(row) > 1 else None
        service       = row[2] if len(row) > 2 else None
        business_line = row[3] if len(row) > 3 else None
        fonction      = row[4] if len(row) > 4 else None
        numero_val    = row[5] if len(row) > 5 else None

        if numero_val is None:
            continue
        numero = str(int(numero_val)) if isinstance(numero_val, (int, float)) else str(numero_val).strip()
        numero = re.sub(r"\D", "", numero)
        if not numero:
            continue

        matricule_s     = str(matricule).strip()     if matricule     not in (None, "") else None
        beneficiaire_s  = str(beneficiaire).strip()   if beneficiaire  not in (None, "") else None
        service_s       = str(service).strip()        if service       not in (None, "") else None
        business_line_s = str(business_line).strip()  if business_line not in (None, "") else None
        fonction_s      = str(fonction).strip()        if fonction      not in (None, "") else None

        sim_obj = db.query(NumeroSIM).filter(NumeroSIM.numero == numero).first()
        if sim_obj:
            if matricule_s:     sim_obj.matricule     = matricule_s
            if beneficiaire_s:  sim_obj.beneficiaire  = beneficiaire_s
            if service_s:       sim_obj.service       = service_s
            if business_line_s: sim_obj.business_line = business_line_s
            if fonction_s:      sim_obj.fonction      = fonction_s
            updated += 1
        else:
            sim_obj = NumeroSIM(
                numero=numero, categorie=CategorieSimEnum.EMPLOYE, statut=StatutSimEnum.ACTIVE,
                matricule=matricule_s, beneficiaire=beneficiaire_s, service=service_s,
                business_line=business_line_s, fonction=fonction_s,
            )
            db.add(sim_obj)
            db.flush()
            created += 1

        # ── Auto-affectation employé ────────────────────────────────────────
        if beneficiaire_s:
            existing_aff = db.query(AffectationSIM).filter(
                AffectationSIM.sim_id == sim_obj.id,
                AffectationSIM.is_active == True,
            ).first()
            if not existing_aff:
                db.add(AffectationSIM(
                    sim_id=sim_obj.id,
                    date_debut=_date.today(),
                    is_active=True,
                    employee_nom=beneficiaire_s,
                    employee_matricule=matricule_s,
                    notes="Importé depuis SUIVI FLOTTE CAMUSAT (ORANGE-mobiles)",
                ))
                affecte += 1

    db.commit()
    return {"created": created, "updated": updated, "affecte": affecte, "errors": errors}


def _import_gps_vehicules_xlsx(content: bytes, db: Session):
    """Import depuis la feuille 'ORANGE-Gps_Vehicules' du fichier de suivi
    flotte (.xlsx) — crée/met à jour les numéros SIM M2M véhicule (forfait),
    les véhicules (modèle / IMEI) et leur affectation. Ignore les lignes
    'Non affecté', les IMEI '#N/A', les lignes de test ('AA TEST...') et la
    ligne 'TOTAL'. La partie facturation (colonnes mensuelles) n'est PAS
    traitée ici."""
    import io, re, unicodedata
    from datetime import date as _date
    from openpyxl import load_workbook

    def norm(s) -> str:
        s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
        return re.sub(r"[^A-Z0-9]", "", s.upper())

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Fichier Excel illisible.")

    created_sim, updated_sim, created_veh, updated_veh, affecte, errors = 0, 0, 0, 0, 0, []

    if "ORANGE-Gps_Vehicules" not in wb.sheetnames:
        return {"created_sim": 0, "updated_sim": 0, "created_vehicule": 0, "updated_vehicule": 0, "affecte": 0, "errors": []}

    ws = wb["ORANGE-Gps_Vehicules"]

    # Repérer la ligne d'en-tête (celle qui contient "N°SIM" / "NSIM")
    header_row_idx = None
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        normalized = [norm(v) if v is not None else "" for v in row]
        if "NSIM" in normalized:
            header_row_idx = r_idx
            break
    if header_row_idx is None:
        return {"created_sim": 0, "updated_sim": 0, "created_vehicule": 0, "updated_vehicule": 0, "affecte": 0,
                "errors": ["Feuille ORANGE-Gps_Vehicules : en-tête introuvable"]}

    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        numero_val = row[0] if len(row) > 0 else None
        if numero_val is None:
            continue
        numero = str(int(numero_val)) if isinstance(numero_val, (int, float)) else str(numero_val).strip()
        numero = re.sub(r"\D", "", numero)
        if not numero:
            continue  # ignore "TOTAL", lignes d'en-tête répétées, etc.

        immat   = row[1] if len(row) > 1 else None
        modele  = row[2] if len(row) > 2 else None
        imei    = row[3] if len(row) > 3 else None
        forfait = row[4] if len(row) > 4 else None

        immat_s  = str(immat).strip()  if immat  not in (None, "") else None
        modele_s = str(modele).strip() if modele not in (None, "") else None
        imei_s   = str(imei).strip()   if imei   not in (None, "") else None
        if imei_s and norm(imei_s) in ("NA", "#NA"):
            imei_s = None
        forfait_v = float(forfait) if isinstance(forfait, (int, float)) else None

        # Lignes de test à ignorer
        if immat_s and norm(immat_s).startswith("AATEST"):
            continue

        sim_obj = db.query(NumeroSIM).filter(NumeroSIM.numero == numero).first()
        if sim_obj:
            if forfait_v is not None:
                sim_obj.forfait = forfait_v
            updated_sim += 1
        else:
            sim_obj = NumeroSIM(
                numero=numero, categorie=CategorieSimEnum.M2M_VEHICULE, statut=StatutSimEnum.ACTIVE,
                forfait=forfait_v,
            )
            db.add(sim_obj)
            db.flush()
            created_sim += 1

        # ── Véhicule + affectation ───────────────────────────────────────────
        if immat_s and norm(immat_s) != "NONAFFECTE":
            veh_obj = db.query(Vehicule).filter(Vehicule.immatriculation == immat_s).first()
            if veh_obj:
                if modele_s: veh_obj.modele = modele_s
                if imei_s:   veh_obj.imei   = imei_s
                updated_veh += 1
            else:
                veh_obj = Vehicule(immatriculation=immat_s, modele=modele_s, imei=imei_s)
                db.add(veh_obj)
                db.flush()
                created_veh += 1

            existing_aff = db.query(AffectationSIM).filter(
                AffectationSIM.sim_id == sim_obj.id,
                AffectationSIM.is_active == True,
            ).first()
            if not existing_aff:
                db.add(AffectationSIM(
                    sim_id=sim_obj.id,
                    vehicule_id=veh_obj.id,
                    date_debut=_date.today(),
                    is_active=True,
                    notes="Importé depuis SUIVI FLOTTE CAMUSAT (ORANGE-Gps_Vehicules)",
                ))
                affecte += 1

    db.commit()
    return {
        "created_sim": created_sim, "updated_sim": updated_sim,
        "created_vehicule": created_veh, "updated_vehicule": updated_veh,
        "affecte": affecte, "errors": errors,
    }


@router.post("/import-global")
async def import_global(file: UploadFile = File(...), db: Session = Depends(get_db), _: User = Depends(require_editor)):
    """Import global depuis le fichier de suivi flotte CAMUSAT (.xlsx unique) :
    remplit automatiquement les numéros SIM employés (feuille
    'ORANGE-mobiles'), les véhicules M2M (feuille 'ORANGE-Gps_Vehicules') et
    les sites RMS (feuilles 'RMS_Orange' / 'RMS_Free'), ainsi que leurs
    affectations respectives.

    La partie facturation (colonnes mensuelles des feuilles, ainsi que
    l'import des factures opérateur) n'est PAS traitée par cette route :
    elle reste gérée séparément via la page Factures.
    """
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "Le fichier doit être au format Excel (.xlsx)")

    content = await file.read()

    result = {
        "employes":  _import_mobiles_xlsx(content, db),
        "vehicules": _import_gps_vehicules_xlsx(content, db),
        "sites":     _import_sites_xlsx(content, db),
    }

    import json
    db.add(ImportGlobalLog(
        nom_fichier=file.filename,
        utilisateur=getattr(_, "full_name", None) or getattr(_, "username", None),
        resultat=json.dumps(result),
    ))
    db.commit()

    return result


@router.get("/import-global/historique", response_model=list[ImportGlobalLogOut])
def historique_import_global(db: Session = Depends(get_db)):
    return (
        db.query(ImportGlobalLog)
        .order_by(ImportGlobalLog.created_at.desc())
        .limit(50)
        .all()
    )


def _relink_lignes_facture(db: Session) -> int:
    """Relie les lignes de factures dont le numéro (numero_raw) correspond
    désormais à un numéro SIM connu mais qui n'avaient pas encore de sim_id
    (ex : la facture a été importée avant que le numéro SIM n'existe).
    Retourne le nombre de lignes mises à jour (commit non inclus)."""
    sim_by_numero = {n: sid for n, sid in db.query(NumeroSIM.numero, NumeroSIM.id).all()}
    if not sim_by_numero:
        return 0

    unmatched = db.query(LigneFacture).filter(LigneFacture.sim_id.is_(None)).all()
    count = 0
    for l in unmatched:
        sid = sim_by_numero.get(l.numero_raw)
        if sid:
            l.sim_id = sid
            l.non_reconnu = "N"
            count += 1
    return count


@router.post("/sites/relink-factures")
def relink_factures(db: Session = Depends(get_db), _: User = Depends(require_editor)):
    """Relance la liaison numéro SIM ↔ lignes de factures (utile après un
    import de sites qui crée de nouveaux numéros SIM)."""
    count = _relink_lignes_facture(db)
    db.commit()
    return {"relinked": count}


@router.patch("/sites/{site_id}", response_model=SiteGSMOut)
def update_site(site_id: int, data: SiteGSMCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(SiteGSM).filter(SiteGSM.id == site_id).first()
    if not obj: raise HTTPException(404, "Site introuvable")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(obj, k, v)
    db.commit(); db.refresh(obj)
    return obj


@router.delete("/sites/{site_id}", status_code=204)
def delete_site(site_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(SiteGSM).filter(SiteGSM.id == site_id).first()
    if not obj: raise HTTPException(404, "Site introuvable")
    db.delete(obj); db.commit()


# ── Véhicules ─────────────────────────────────────────────────────────────────

def _vehicule_sim_map(db: Session) -> dict:
    """Retourne un dict {vehicule_id: (sim_numero, sim_id, operateur, statut, forfait)}
    pour toutes les affectations actives."""
    rows = (
        db.query(AffectationSIM.vehicule_id, NumeroSIM.numero, NumeroSIM.id,
                 NumeroSIM.operateur, NumeroSIM.statut, NumeroSIM.forfait)
        .join(NumeroSIM, NumeroSIM.id == AffectationSIM.sim_id)
        .filter(
            AffectationSIM.is_active == True,
            AffectationSIM.vehicule_id.isnot(None),
        )
        .all()
    )
    return {r.vehicule_id: (r.numero, r.id, r.operateur, r.statut, r.forfait) for r in rows}


@router.get("/vehicules", response_model=list[VehiculeOut])
def list_vehicules(db: Session = Depends(get_db)):
    vehicules = db.query(Vehicule).order_by(Vehicule.immatriculation).all()
    sim_map   = _vehicule_sim_map(db)
    fact_map  = _latest_facture_map(db, type_ligne="GPS")
    result = []
    for v in vehicules:
        item = VehiculeOut.model_validate(v)
        sim_info = sim_map.get(v.id)
        if sim_info:
            item.sim_numero    = sim_info[0]
            item.sim_operateur = sim_info[2]
            item.statut_sim    = sim_info[3].value if sim_info[3] else None
            item.forfait       = float(sim_info[4]) if sim_info[4] is not None else None
            item.derniere_facture = fact_map.get(sim_info[1])
        result.append(item)
    return result


@router.get("/vehicules/export")
def export_vehicules(db: Session = Depends(get_db)):
    vehicules = db.query(Vehicule).order_by(Vehicule.immatriculation).all()
    sim_map   = _vehicule_sim_map(db)
    rows = [["Immatriculation", "Marque", "Modèle", "Affectation", "Numéro SIM"]]
    for v in vehicules:
        rows.append([
            v.immatriculation,
            v.marque      or "",
            v.modele      or "",
            v.affectation or "",
            (sim_map.get(v.id) or (None,))[0] or "",
        ])
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerows(rows)
    content = chr(0xFEFF) + output.getvalue()
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8")),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=vehicules.csv"},
    )


@router.get("/vehicules/export-excel")
def export_vehicules_excel(
    filter_sim:  Optional[str]  = None,
    search:      Optional[str]  = None,
    cols:        Optional[str]  = None,
    date_debut:  Optional[date] = None,
    date_fin:    Optional[date] = None,
    db: Session = Depends(get_db),
):
    """Export Excel stylisé des véhicules."""
    import io as _io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    vehicules = db.query(Vehicule).order_by(Vehicule.immatriculation).all()
    sim_map   = _vehicule_sim_map(db)

    fact_map = _latest_facture_map(db, type_ligne="GPS")

    class VRow:
        def __init__(self, v):
            self.id = v.id; self.immatriculation = v.immatriculation
            self.marque = v.marque; self.modele = v.modele; self.imei = v.imei
            self.affectation = v.affectation; self.created_at = v.created_at
            sim_info = sim_map.get(v.id)
            if sim_info:
                self.sim_numero = sim_info[0]
                self.forfait    = float(sim_info[4]) if sim_info[4] is not None else None
                self.derniere_facture = fact_map.get(sim_info[1])
            else:
                self.sim_numero = None
                self.forfait = None
                self.derniere_facture = None

    rows_data = [VRow(v) for v in vehicules]
    if filter_sim == "affecte":      rows_data = [r for r in rows_data if r.sim_numero]
    elif filter_sim == "non_affecte": rows_data = [r for r in rows_data if not r.sim_numero]
    if date_debut: rows_data = [r for r in rows_data if r.created_at and r.created_at.date() >= date_debut]
    if date_fin:   rows_data = [r for r in rows_data if r.created_at and r.created_at.date() <= date_fin]
    if search:
        q = search.lower()
        rows_data = [r for r in rows_data if
            q in (r.immatriculation or "").lower() or q in (r.marque or "").lower() or
            q in (r.modele or "").lower() or q in (r.sim_numero or "").lower()]

    ALL_COLS = [
        ("immat",      "Immatriculation",  lambda r: r.immatriculation or ""),
        ("marque",     "Marque",           lambda r: r.marque or ""),
        ("modele",     "Modèle",           lambda r: r.modele or ""),
        ("affectation","Affectation",      lambda r: r.affectation or ""),
        ("sim_numero", "Numéro SIM",       lambda r: r.sim_numero or ""),
        ("imei",       "IMEI",             lambda r: r.imei or ""),
        ("forfait",    "Forfait",          lambda r: r.forfait if r.forfait is not None else ""),
        ("statut_sim", "Statut SIM",       lambda r: "Affecté" if r.sim_numero else "Non affecté"),
        ("derniere_facture", "Dernière facture", lambda r: r.derniere_facture["montant_ttc"] if r.derniere_facture and r.derniere_facture.get("montant_ttc") is not None else ""),
        ("created_at", "Date création",    lambda r: r.created_at.strftime("%d/%m/%Y") if r.created_at else ""),
    ]
    selected_keys = set(cols.split(",")) if cols else {c[0] for c in ALL_COLS}
    columns = [(k, lbl, fn) for k, lbl, fn in ALL_COLS if k in selected_keys]

    BLUE_HDR = "1B3D6F"; WHITE = "FFFFFF"; ROW_EVEN = "EEF4FF"; BORDER_COL = "C5D3E8"
    thin = Side(style="thin", color=BORDER_COL)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook(); ws = wb.active; ws.title = "Véhicules"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    tc = ws.cell(row=1, column=1, value="INVENTAIRE DES VÉHICULES")
    tc.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=BLUE_HDR)
    tc.alignment = Alignment(horizontal="center", vertical="center"); ws.row_dimensions[1].height = 32

    parts = [f"Exporté le {datetime.now().strftime('%d/%m/%Y')}"]
    if date_debut: parts.append(f"Du {date_debut.strftime('%d/%m/%Y')}")
    if date_fin:   parts.append(f"au {date_fin.strftime('%d/%m/%Y')}")
    if filter_sim == "affecte":       parts.append("Avec SIM affecté")
    elif filter_sim == "non_affecte": parts.append("Sans SIM affecté")
    parts.append(f"{len(rows_data)} véhicule(s)")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sc = ws.cell(row=2, column=1, value="  |  ".join(parts))
    sc.font = Font(name="Calibri", italic=True, size=9, color="5B7DB1")
    sc.fill = PatternFill("solid", fgColor="D9E6F7")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1); ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    HDR_ROW = 4
    for ci, (_, lbl, _fn) in enumerate(columns, start=1):
        c = ws.cell(row=HDR_ROW, column=ci, value=lbl)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = border_all
    ws.row_dimensions[HDR_ROW].height = 24; ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    for ri, r in enumerate(rows_data):
        row_num = HDR_ROW + 1 + ri
        row_fill = PatternFill("solid", fgColor=(ROW_EVEN if ri % 2 == 0 else WHITE))
        for ci, (_, _lbl, fn) in enumerate(columns, start=1):
            c = ws.cell(row=row_num, column=ci, value=fn(r))
            c.fill = row_fill; c.font = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", horizontal="left", indent=1); c.border = border_all
        ws.row_dimensions[row_num].height = 18

    for ci, (_, lbl, fn) in enumerate(columns, start=1):
        max_len = max((len(fn(r)) for r in rows_data), default=0)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(lbl) + 2, min(max_len + 2, 40))

    ws.oddFooter.center.text = "&\"Calibri\"&8 CAMUSAT — Véhicules  |  Page &P / &N"
    buf = _io.BytesIO(); wb.save(buf); buf.seek(0)

    fname = "vehicules"
    if date_debut: fname += f"_du_{date_debut}"
    if date_fin:   fname += f"_au_{date_fin}"
    fname += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/vehicules/import")
async def import_vehicules(file: UploadFile = File(...), db: Session = Depends(get_db), _: User = Depends(require_editor)):
    """
    Formats acceptés (séparateur ; ou tabulation) :
    - Nouveau format CAMUSAT : Numéro ; IMSI ; IMMATRICULATION ; MODEL
    - Ancien format           : Immatriculation ; Marque ; Modèle ; Affectation

    Si la colonne "Numéro" contient un numéro SIM existant,
    une affectation est créée automatiquement entre ce SIM et le véhicule.
    """
    import unicodedata as _ud
    from datetime import date as _date

    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    def norm(s: str) -> str:
        s = _ud.normalize("NFD", s)
        s = "".join(c for c in s if _ud.category(c) != "Mn")
        return s.strip().upper().replace(" ", "_").replace("/", "_").replace("'", "")

    first_line = text.split("\n")[0]
    sep = ";" if ";" in first_line else "\t"

    reader = csv.DictReader(io.StringIO(text), delimiter=sep)
    created, updated, affecte, errors = 0, 0, 0, []

    for i, raw_row in enumerate(reader, start=2):
        row = {norm(k): (v.strip() if v else "") for k, v in raw_row.items()}

        # Immatriculation
        immat = (row.get("IMMATRICULATION") or row.get("IMMAT") or "").strip()
        if not immat:
            errors.append({"ligne": i, "message": "Immatriculation manquante"}); continue

        # Modèle (nouveau : MODEL, ancien : MODELE)
        modele      = row.get("MODEL") or row.get("MODELE") or row.get("MODELE_") or None
        marque      = row.get("MARQUE") or None
        affectation = row.get("AFFECTATION") or None

        # Numéro SIM (nouveau format)
        sim_numero = (row.get("NUMERO") or row.get("NUM") or "").strip() or None

        # Créer ou mettre à jour
        existing = db.query(Vehicule).filter(Vehicule.immatriculation == immat).first()
        if existing:
            if modele:      existing.modele      = modele
            if marque:      existing.marque      = marque
            if affectation: existing.affectation = affectation
            veh_obj = existing
            updated += 1
        else:
            veh_obj = Vehicule(immatriculation=immat, marque=marque,
                               modele=modele, affectation=affectation)
            db.add(veh_obj)
            db.flush()
            created += 1

        # Auto-affectation SIM
        if sim_numero:
            sim_obj = db.query(NumeroSIM).filter(NumeroSIM.numero == sim_numero).first()
            if sim_obj:
                existing_aff = db.query(AffectationSIM).filter(
                    AffectationSIM.vehicule_id == veh_obj.id,
                    AffectationSIM.is_active == True,
                ).first()
                if not existing_aff:
                    old_aff = db.query(AffectationSIM).filter(
                        AffectationSIM.sim_id == sim_obj.id,
                        AffectationSIM.is_active == True,
                    ).first()
                    if old_aff:
                        old_aff.is_active = False
                        old_aff.date_fin  = _date.today()
                    db.add(AffectationSIM(
                        sim_id      = sim_obj.id,
                        vehicule_id = veh_obj.id,
                        date_debut  = _date.today(),
                        is_active   = True,
                    ))
                    affecte += 1

    db.commit()
    return {"created": created, "updated": updated,
            "affecte": affecte, "errors": errors,
            "total_lignes": created + updated + len(errors)}


@router.post("/vehicules", response_model=VehiculeOut, status_code=201)
def create_vehicule(data: VehiculeCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    if db.query(Vehicule).filter(Vehicule.immatriculation == data.immatriculation).first():
        raise HTTPException(400, "Immatriculation déjà enregistrée")
    obj = Vehicule(**data.model_dump())
    db.add(obj); db.commit(); db.refresh(obj)
    return obj


@router.patch("/vehicules/{vehicule_id}", response_model=VehiculeOut)
def update_vehicule(vehicule_id: int, data: VehiculeCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Vehicule).filter(Vehicule.id == vehicule_id).first()
    if not obj: raise HTTPException(404, "Véhicule introuvable")
    for k, v in data.model_dump(exclude_none=True).items():
        setattr(obj, k, v)
    db.commit(); db.refresh(obj)
    return obj


@router.delete("/vehicules/{vehicule_id}", status_code=204)
def delete_vehicule(vehicule_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Vehicule).filter(Vehicule.id == vehicule_id).first()
    if not obj: raise HTTPException(404, "Véhicule introuvable")
    db.delete(obj); db.commit()
