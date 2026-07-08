import io
import json
from datetime import datetime
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import Optional

from ..database import get_db
from ..models.materiel import Materiel, StatutMateriel
from ..models.attribution import Attribution, StatutAttribution
from ..models.telephonie import NumeroSIM, SiteGSM, Vehicule, AffectationSIM
from ..models.facture import FactureTelecom
from ..models.export_log import ExportLog
from ..models.user import User
from ..services.auth_service import require_editor

router = APIRouter(prefix="/export-global", tags=["Export Global"])

BLUE_HDR   = "1B3D6F"
WHITE      = "FFFFFF"
ROW_EVEN   = "EEF4FF"
ROW_ODD    = "FFFFFF"
BORDER_COL = "C5D3E8"


def _style_sheet(ws, columns: list, rows: list):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin   = Side(style="thin", color=BORDER_COL)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, col in enumerate(columns, start=1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = border
    ws.row_dimensions[1].height = 22

    for ri, row in enumerate(rows, start=2):
        fill_color = ROW_EVEN if ri % 2 == 0 else ROW_ODD
        for ci, val in enumerate(row, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = PatternFill("solid", fgColor=fill_color)
            c.font      = Font(name="Calibri", size=9)
            c.alignment = Alignment(vertical="center")
            c.border    = border

    for ci, col in enumerate(columns, start=1):
        max_w = max(
            len(str(col)),
            max((len(str(r[ci - 1] or "")) for r in rows), default=0),
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 4, 40)

    ws.freeze_panes = "A2"


# ─── GET /export-global/logs ─────────────────────────────────────────────────
@router.get("/logs", summary="Historique des exports globaux")
def get_export_logs(
    limit: int = Query(50, ge=1, le=200),
    db:    Session = Depends(get_db),
    _:     User    = Depends(require_editor),
):
    logs = (
        db.query(ExportLog)
        .order_by(ExportLog.created_at.desc())
        .limit(limit)
        .all()
    )
    return [
        {
            "id":         log.id,
            "user_name":  log.user_name,
            "filename":   log.filename,
            "filters":    json.loads(log.filters)  if log.filters  else {},
            "nb_rows":    json.loads(log.nb_rows)  if log.nb_rows  else {},
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]


# ─── GET /export-global ───────────────────────────────────────────────────────
@router.get("", summary="Export global — un seul fichier multi-feuilles")
def export_global(
    mat_statut:      Optional[str] = Query(None),
    mat_type:        Optional[str] = Query(None),
    att_statut:      Optional[str] = Query(None),
    sim_categorie:   Optional[str] = Query(None),
    sim_statut:      Optional[str] = Query(None),
    site_filter_sim: Optional[str] = Query(None),
    veh_filter_sim:  Optional[str] = Query(None),
    fact_annee:      Optional[int] = Query(None),
    fact_mois:       Optional[int] = Query(None),
    db:              Session = Depends(get_db),
    current_user:    User    = Depends(require_editor),
):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Matériels ─────────────────────────────────────────────────────────
    q = db.query(Materiel).options(joinedload(Materiel.attributions))
    if mat_statut: q = q.filter(Materiel.statut == mat_statut)
    if mat_type:   q = q.filter(Materiel.type_materiel == mat_type)
    materiels = q.order_by(Materiel.marque).all()

    def active_att(m):
        return next((a for a in m.attributions if a.statut == StatutAttribution.ACTIVE), None)

    TYPE_LBL = {
        "ORDINATEUR_PORTABLE": "PC Portable", "ORDINATEUR_FIXE": "PC Fixe",
        "ECRAN": "Écran", "SOURIS": "Souris", "CLAVIER": "Clavier",
        "TELEPHONE": "Téléphone", "IMPRIMANTE": "Imprimante",
        "SWITCH": "Switch", "ROUTEUR": "Routeur", "ONDULEUR": "Onduleur", "AUTRE": "Autre",
        "TABLETTE": "Tablette", "SERVEUR": "Serveur", "AP": "AP", "PARE_FEU": "Pare-feu",
    }
    STATUT_MAT = {
        "DISPONIBLE": "Disponible", "ATTRIBUE": "Attribué",
        "MAINTENANCE": "Maintenance", "EN_PANNE": "En panne", "REFORME": "Réformé",
    }
    ws_mat = wb.create_sheet("Matériels")
    mat_cols = ["ID", "Type", "Marque", "Modèle", "N° Série", "Adresse MAC",
                "Référence", "État", "Statut", "Date Acquisition", "Assigné à"]
    mat_rows = []
    for m in materiels:
        a = active_att(m)
        mat_rows.append([
            m.id,
            TYPE_LBL.get(m.type_materiel or "", m.type_materiel or ""),
            m.marque or "",
            m.modele or "",
            m.numero_serie or "",
            m.adresse_mac or "",
            m.reference or "",
            m.etat or "",
            STATUT_MAT.get(m.statut or "", m.statut or ""),
            m.date_acquisition.strftime("%d/%m/%Y") if m.date_acquisition else "",
            f"{a.employee_prenom or ''} {a.employee_nom}".strip() if a else "",
        ])
    _style_sheet(ws_mat, mat_cols, mat_rows)

    # ── 2. Attributions ───────────────────────────────────────────────────────
    q_att = db.query(Attribution).options(joinedload(Attribution.materiel))
    if att_statut: q_att = q_att.filter(Attribution.statut == att_statut)
    attributions = q_att.order_by(Attribution.date_attribution.desc()).all()

    ws_att = wb.create_sheet("Attributions")
    att_cols = ["ID", "Matériel", "N° Série", "Employé", "Matricule",
                "Direction", "Date Attribution", "Date Retour", "Statut"]
    att_rows = []
    for att in attributions:
        mat = att.materiel
        att_rows.append([
            att.id,
            f"{mat.marque or ''} {mat.modele or ''}".strip() if mat else "",
            mat.numero_serie or "" if mat else "",
            f"{att.employee_prenom or ''} {att.employee_nom or ''}".strip(),
            att.employee_matricule or "",
            att.employee_direction or "",
            att.date_attribution.strftime("%d/%m/%Y") if att.date_attribution else "",
            att.date_retour.strftime("%d/%m/%Y") if att.date_retour else "",
            att.statut or "",
        ])
    _style_sheet(ws_att, att_cols, att_rows)

    # ── 3. Numéros SIM ────────────────────────────────────────────────────────
    q_sim = db.query(NumeroSIM).options(
        joinedload(NumeroSIM.affectations).joinedload(AffectationSIM.site),
        joinedload(NumeroSIM.affectations).joinedload(AffectationSIM.vehicule),
    )
    if sim_categorie: q_sim = q_sim.filter(NumeroSIM.categorie == sim_categorie)
    if sim_statut:    q_sim = q_sim.filter(NumeroSIM.statut == sim_statut)
    sims = q_sim.order_by(NumeroSIM.numero).all()

    def active_aff(sim):
        return next((a for a in sim.affectations if a.is_active), None)

    ws_sim = wb.create_sheet("Numéros SIM")
    sim_cols = ["ID", "Numéro", "Catégorie", "Opérateur", "Statut",
                "Forfait", "Date Activation", "Affecté à"]
    sim_rows = []
    for s in sims:
        aff   = active_aff(s)
        affecte = ""
        if aff:
            if aff.site_id and aff.site:
                affecte = aff.site.nom or f"Site #{aff.site_id}"
            elif aff.vehicule_id and aff.vehicule:
                affecte = aff.vehicule.immatriculation or f"Véhicule #{aff.vehicule_id}"
            elif aff.employee_nom:
                affecte = f"{aff.employee_nom} ({aff.employee_matricule or ''})"
        sim_rows.append([
            s.id, s.numero or "", s.categorie or "", s.operateur or "",
            s.statut or "", s.forfait or "",
            s.date_activation.strftime("%d/%m/%Y") if s.date_activation else "",
            affecte,
        ])
    _style_sheet(ws_sim, sim_cols, sim_rows)

    # ── 4. Sites RMS ──────────────────────────────────────────────────────────
    q_site = db.query(SiteGSM).options(
        joinedload(SiteGSM.affectations).joinedload(AffectationSIM.sim)
    )
    sites = q_site.order_by(SiteGSM.nom).all()

    def active_sim_site(site):
        for aff in site.affectations:
            if aff.is_active and aff.sim:
                return aff.sim.numero
        return ""

    if site_filter_sim == "with_sim":
        sites = [s for s in sites if active_sim_site(s)]
    elif site_filter_sim == "without_sim":
        sites = [s for s in sites if not active_sim_site(s)]

    ws_site = wb.create_sheet("Sites RMS")
    site_cols = ["ID", "Code Site", "Nom", "Région", "Latitude", "Longitude", "SIM active"]
    site_rows = [[s.id, s.code_site or "", s.nom or "", s.region or "",
                  s.latitude or "", s.longitude or "", active_sim_site(s)] for s in sites]
    _style_sheet(ws_site, site_cols, site_rows)

    # ── 5. Véhicules GPS ──────────────────────────────────────────────────────
    q_veh = db.query(Vehicule).options(
        joinedload(Vehicule.affectations).joinedload(AffectationSIM.sim)
    )
    vehicules = q_veh.order_by(Vehicule.immatriculation).all()

    def active_sim_veh(v):
        for aff in v.affectations:
            if aff.is_active and aff.sim:
                return aff.sim.numero
        return ""

    if veh_filter_sim == "with_sim":
        vehicules = [v for v in vehicules if active_sim_veh(v)]
    elif veh_filter_sim == "without_sim":
        vehicules = [v for v in vehicules if not active_sim_veh(v)]

    ws_veh = wb.create_sheet("Véhicules GPS")
    veh_cols = ["ID", "Immatriculation", "Marque", "Modèle", "Type", "IMEI", "SIM active"]
    veh_rows = [[v.id, v.immatriculation or "", v.marque or "", v.modele or "",
                 v.type_vehicule or "", v.imei or "", active_sim_veh(v)] for v in vehicules]
    _style_sheet(ws_veh, veh_cols, veh_rows)

    # ── 6. Factures ───────────────────────────────────────────────────────────
    q_fact = db.query(FactureTelecom)
    if fact_annee: q_fact = q_fact.filter(FactureTelecom.annee == fact_annee)
    if fact_mois:  q_fact = q_fact.filter(FactureTelecom.mois  == fact_mois)
    factures = q_fact.order_by(FactureTelecom.annee.desc(), FactureTelecom.mois.desc()).all()

    ws_fact = wb.create_sheet("Factures Télécom")
    fact_cols = ["ID", "Période", "Opérateur", "N° Facture",
                 "Montant HT", "Montant TTC", "Solde Facture", "Statut Paiement"]
    fact_rows = [[
        f.id,
        f"{f.mois:02d}/{f.annee}" if f.mois and f.annee else "",
        f.operateur or "", f.numero_facture or "",
        float(f.montant_ht)    if f.montant_ht    else "",
        float(f.montant_ttc)   if f.montant_ttc   else "",
        float(f.solde_facture) if f.solde_facture else "",
        f.statut_paiement or "",
    ] for f in factures]
    _style_sheet(ws_fact, fact_cols, fact_rows)

    # ── Page de garde ─────────────────────────────────────────────────────────
    nb_rows = {
        "Matériels":    len(mat_rows),
        "Attributions": len(att_rows),
        "Numéros SIM":  len(sim_rows),
        "Sites RMS":    len(site_rows),
        "Véhicules GPS":len(veh_rows),
        "Factures":     len(fact_rows),
    }
    ws_cover = wb.create_sheet("Résumé", 0)
    ws_cover.column_dimensions["A"].width = 30
    ws_cover.column_dimensions["B"].width = 20

    cover_data = [
        ("Export Global — Camusat Sénégal", None),
        ("", None),
        ("Généré le",       datetime.now().strftime("%d/%m/%Y à %Hh%M")),
        ("Généré par",      current_user.full_name or current_user.username),
        ("", None),
        ("Feuille", "Nombre de lignes"),
        *[(k, v) for k, v in nb_rows.items()],
    ]
    for ri, (a, b) in enumerate(cover_data, start=1):
        ca = ws_cover.cell(row=ri, column=1, value=a)
        if ri == 1:
            ws_cover.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ca.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
            ca.fill      = PatternFill("solid", fgColor=BLUE_HDR)
            ca.alignment = Alignment(horizontal="center", vertical="center")
            ws_cover.row_dimensions[1].height = 32
        elif ri == 6:
            for ci in (1, 2):
                c = ws_cover.cell(row=ri, column=ci, value=(a if ci == 1 else b))
                c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
                c.fill = PatternFill("solid", fgColor=BLUE_HDR)
        elif ri > 6:
            fill_c = ROW_EVEN if ri % 2 == 0 else ROW_ODD
            for ci, val in ((1, a), (2, b)):
                c = ws_cover.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Calibri", size=10)
                c.fill = PatternFill("solid", fgColor=fill_c)
        else:
            ca.font = Font(name="Calibri", size=10)
            if b is not None:
                ws_cover.cell(row=ri, column=2, value=b).font = Font(name="Calibri", size=10)

    # ── Enregistrement du log ─────────────────────────────────────────────────
    active_filters = {k: v for k, v in {
        "Statut matériel": mat_statut,  "Type matériel": mat_type,
        "Statut attribution": att_statut,
        "Catégorie SIM": sim_categorie, "Statut SIM": sim_statut,
        "Sites SIM": site_filter_sim,   "Véhicules SIM": veh_filter_sim,
        "Année facture": fact_annee,    "Mois facture": fact_mois,
    }.items() if v}

    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"export_global_{date_str}.xlsx"

    db.add(ExportLog(
        user_name=current_user.full_name or current_user.username,
        filename=filename,
        filters=json.dumps(active_filters, ensure_ascii=False),
        nb_rows=json.dumps(nb_rows, ensure_ascii=False),
    ))
    db.commit()

    # ── Réponse ───────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
