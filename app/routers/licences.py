from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import Optional
from datetime import date, datetime
from pydantic import BaseModel
import csv, io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from ..database import get_db
from ..models.licence import Licence, LicenceAttribution
from ..models.materiel import Materiel
from ..models.user import User
from ..services.auth_service import require_editor

router = APIRouter(prefix="/api/licences", tags=["Licences"])


class LicenceIn(BaseModel):
    logiciel:        str
    editeur:         Optional[str]  = None
    version:         Optional[str]  = None
    cle_licence:     Optional[str]  = None
    date_achat:      Optional[date] = None
    date_expiration: Optional[date] = None
    nb_postes_max:   Optional[int]  = None
    notes:           Optional[str]  = None


class LicencePatch(BaseModel):
    logiciel:        Optional[str]  = None
    editeur:         Optional[str]  = None
    version:         Optional[str]  = None
    cle_licence:     Optional[str]  = None
    date_achat:      Optional[date] = None
    date_expiration: Optional[date] = None
    nb_postes_max:   Optional[int]  = None
    notes:           Optional[str]  = None


class AttributionIn(BaseModel):
    employee_nom:       str
    employee_prenom:    Optional[str] = None
    employee_matricule: Optional[str] = None
    employee_service:   Optional[str] = None
    materiel_id:        Optional[int] = None
    date_attribution:   date
    notes:              Optional[str] = None


# ── Recherche employés (depuis beneficiaire_* des matériels) ─────────────────

@router.get("/employes-search")
def search_employes(q: str = Query(""), db: Session = Depends(get_db)):
    rows = (
        db.query(
            Materiel.beneficiaire_nom,
            Materiel.beneficiaire_prenom,
            Materiel.beneficiaire_matricule,
        )
        .filter(Materiel.beneficiaire_nom.isnot(None))
        .group_by(
            Materiel.beneficiaire_nom,
            Materiel.beneficiaire_prenom,
            Materiel.beneficiaire_matricule,
        )
        .all()
    )
    seen: set[str] = set()
    result = []
    for nom, prenom, mat in rows:
        key = f"{(nom or '').strip().lower()}|{(prenom or '').strip().lower()}"
        if key in seen:
            continue
        seen.add(key)
        full = f"{(nom or '').strip()} {(prenom or '').strip()}".strip()
        if q and q.lower() not in full.lower() and q.lower() not in (mat or "").lower():
            continue
        result.append({
            "nom":       (nom or "").strip(),
            "prenom":    (prenom or "").strip(),
            "matricule": (mat or "").strip(),
        })
    result.sort(key=lambda x: x["nom"])
    return result[:30]


# ── Helpers Excel ─────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

COLS = [
    ("logiciel",        "Logiciel",         30),
    ("cle_licence",     "Clé de licence",   35),
    ("date_achat",      "Date d'achat",     15),
    ("date_expiration", "Date d'expiration",18),
]

COLS_EXPORT = COLS + [
    ("employee_prenom",    "Prénom",              18),
    ("employee_nom",       "Nom",                 18),
    ("employee_matricule", "Matricule",           14),
    ("employee_service",   "Service",             20),
    ("date_attribution",   "Date d'attribution",  18),
]


def _make_header(ws, cols=None):
    if cols is None:
        cols = COLS
    for col_idx, (_, label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill  = _HEADER_FILL
        cell.font  = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    ws.row_dimensions[1].height = 22


def _xlsx_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _parse_date(s):
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


# ── Télécharger modèle Excel ──────────────────────────────────────────────────

@router.get("/template")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modèle"
    _make_header(ws)
    # Ligne exemple
    ws.append(["Microsoft Office 365", "XXXXX-XXXXX-XXXXX-XXXXX", "2024-01-01", "2025-01-01"])
    return _xlsx_response(wb, "modele_licences.xlsx")


# ── Exporter données réelles ───────────────────────────────────────────────────

@router.get("/export")
def export_licences(
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    from openpyxl.styles import Border, Side, Alignment as Align
    from openpyxl.utils import get_column_letter

    q = db.query(Licence).options(joinedload(Licence.attributions))
    if search:
        q = q.filter(Licence.logiciel.ilike(f"%{search}%"))
    licences = q.order_by(Licence.logiciel.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Licences"
    _make_header(ws, COLS_EXPORT)

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    grey_fill = PatternFill("solid", fgColor="F5F5F5")

    row_idx = 2
    for l in licences:
        lic_vals = [
            l.logiciel,
            l.cle_licence or "",
            l.date_achat.isoformat() if l.date_achat else "",
            l.date_expiration.isoformat() if l.date_expiration else "",
        ]
        atts = l.attributions or []
        n = max(len(atts), 1)

        for i, att in enumerate(atts if atts else [None]):
            att_vals = [
                att.employee_prenom or "" if att else "",
                att.employee_nom    or "" if att else "",
                att.employee_matricule or "" if att else "",
                att.employee_service   or "" if att else "",
                att.date_attribution.isoformat() if att and att.date_attribution else "",
            ]
            ws.append(lic_vals + att_vals if i == 0 else ["", "", "", ""] + att_vals)
            for col in range(1, len(COLS_EXPORT) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border
                if row_idx % 2 == 0:
                    cell.fill = grey_fill
            row_idx += 1

        # Fusionner les colonnes licence si plusieurs attributions
        if n > 1:
            for col in range(1, len(COLS) + 1):
                ws.merge_cells(
                    start_row=row_idx - n, end_row=row_idx - 1,
                    start_column=col, end_column=col,
                )
                ws.cell(row=row_idx - n, column=col).alignment = Align(vertical="center", wrap_text=True)

    return _xlsx_response(wb, "licences.xlsx")


# ── Import Excel / CSV ────────────────────────────────────────────────────────

@router.post("/import", status_code=201)
async def import_licences(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    content = await file.read()
    created = 0
    errors  = []

    if file.filename and file.filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col_map = {h: i for i, h in enumerate(headers)}
        def cell(row, key):
            idx = col_map.get(key)
            return str(row[idx].value or "").strip() if idx is not None else ""
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            logiciel = cell(row, "logiciel")
            if not logiciel:
                errors.append(f"Ligne {row_idx} : logiciel manquant"); continue
            db.add(Licence(
                logiciel=logiciel,
                cle_licence=cell(row, "clé de licence") or cell(row, "cle_licence") or None,
                date_achat=_parse_date(cell(row, "date d'achat") or cell(row, "date_achat")),
                date_expiration=_parse_date(cell(row, "date d'expiration") or cell(row, "date_expiration")),
            ))
            created += 1
    else:
        try:    text = content.decode("utf-8-sig")
        except: text = content.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        for i, row in enumerate(reader, start=2):
            logiciel = (row.get("logiciel") or "").strip()
            if not logiciel:
                errors.append(f"Ligne {i} : logiciel manquant"); continue
            db.add(Licence(
                logiciel=logiciel,
                cle_licence=(row.get("cle_licence") or "").strip() or None,
                date_achat=_parse_date(row.get("date_achat")),
                date_expiration=_parse_date(row.get("date_expiration")),
            ))
            created += 1

    db.commit()
    return {"created": created, "errors": errors}


# ── Licences ──────────────────────────────────────────────────────────────────

@router.get("/")
def list_licences(
    search: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    q = db.query(Licence).options(joinedload(Licence.attributions))
    if search:
        q = q.filter(Licence.logiciel.ilike(f"%{search}%"))
    return q.order_by(Licence.logiciel.asc()).all()


@router.post("/", status_code=201)
def create_licence(data: LicenceIn, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = Licence(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.patch("/{licence_id}")
def update_licence(licence_id: int, data: LicencePatch, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Licence).filter(Licence.id == licence_id).first()
    if not obj:
        raise HTTPException(404, "Licence introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/{licence_id}", status_code=204)
def delete_licence(licence_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Licence).filter(Licence.id == licence_id).first()
    if not obj:
        raise HTTPException(404, "Licence introuvable")
    db.delete(obj)
    db.commit()


# ── Attributions de licence ───────────────────────────────────────────────────

@router.post("/{licence_id}/attributions", status_code=201)
def add_attribution(licence_id: int, data: AttributionIn, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    if not db.query(Licence).filter(Licence.id == licence_id).first():
        raise HTTPException(404, "Licence introuvable")
    obj = LicenceAttribution(licence_id=licence_id, **data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/{licence_id}/attributions/{attr_id}", status_code=204)
def remove_attribution(licence_id: int, attr_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(LicenceAttribution).filter(
        LicenceAttribution.id == attr_id,
        LicenceAttribution.licence_id == licence_id,
    ).first()
    if not obj:
        raise HTTPException(404, "Attribution introuvable")
    db.delete(obj)
    db.commit()
