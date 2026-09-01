from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_
from typing import Optional
from datetime import date
from pydantic import BaseModel
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ..database import get_db
from sqlalchemy.orm import joinedload
from ..models.planning import Tache, StatutTache, PrioriteTache, Checklist, ChecklistItem
from ..models.user import User
from ..services.auth_service import require_editor

router = APIRouter(prefix="/api/planning", tags=["Planning"])


class TacheIn(BaseModel):
    titre:          str
    description:    Optional[str]  = None
    date_planifiee: date
    date_fin:       Optional[date] = None
    statut:         StatutTache    = StatutTache.A_FAIRE
    priorite:       PrioriteTache  = PrioriteTache.NORMALE
    responsable:    Optional[str]  = None
    notes:          Optional[str]  = None


class TachePatch(BaseModel):
    titre:          Optional[str]          = None
    description:    Optional[str]          = None
    date_planifiee: Optional[date]         = None
    date_fin:       Optional[date]         = None
    statut:         Optional[StatutTache]  = None
    priorite:       Optional[PrioriteTache] = None
    responsable:    Optional[str]          = None
    notes:          Optional[str]          = None


@router.get("/")
def list_taches(
    date_debut:  Optional[date] = Query(None),
    date_fin:    Optional[date] = Query(None),
    statut:      Optional[str]  = Query(None),
    responsable: Optional[str]  = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    q = db.query(Tache)
    if date_debut:
        q = q.filter(Tache.date_planifiee >= date_debut)
    if date_fin:
        q = q.filter(Tache.date_planifiee <= date_fin)
    if statut:
        q = q.filter(Tache.statut == statut)
    if responsable:
        q = q.filter(Tache.responsable.ilike(f"%{responsable}%"))
    return q.order_by(Tache.date_planifiee.asc(), Tache.priorite.desc()).all()


@router.post("/", status_code=201)
def create_tache(data: TacheIn, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = Tache(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.patch("/{tache_id}")
def update_tache(tache_id: int, data: TachePatch, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Tache).filter(Tache.id == tache_id).first()
    if not obj:
        raise HTTPException(404, "Tâche introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/{tache_id}", status_code=204)
def delete_tache(tache_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Tache).filter(Tache.id == tache_id).first()
    if not obj:
        raise HTTPException(404, "Tâche introuvable")
    db.delete(obj)
    db.commit()


# ── Checklists ────────────────────────────────────────────────────────────────

class ChecklistItemIn(BaseModel):
    titre: str
    ordre: int = 0

class ChecklistIn(BaseModel):
    nom:         Optional[str] = None
    description: Optional[str] = None
    responsable: Optional[str] = None
    items:       list[ChecklistItemIn] = []

class ChecklistPatch(BaseModel):
    nom:         Optional[str] = None
    description: Optional[str] = None
    responsable: Optional[str] = None

class ItemPatch(BaseModel):
    titre:  Optional[str]  = None
    cochee: Optional[bool] = None
    ordre:  Optional[int]  = None


@router.get("/checklists")
def list_checklists(db: Session = Depends(get_db), _: User = Depends(require_editor)):
    return db.query(Checklist).options(joinedload(Checklist.items)).order_by(Checklist.created_at.desc()).all()


@router.post("/checklists", status_code=201)
def create_checklist(data: ChecklistIn, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    from datetime import datetime
    nom = data.nom or f"Checklist du {datetime.now().strftime('%d/%m/%Y')}"
    cl = Checklist(nom=nom, description=data.description, responsable=data.responsable)
    db.add(cl)
    db.flush()
    for i, item in enumerate(data.items):
        db.add(ChecklistItem(checklist_id=cl.id, titre=item.titre, ordre=i))
    db.commit()
    db.refresh(cl)
    return cl


@router.patch("/checklists/{cl_id}")
def update_checklist(cl_id: int, data: ChecklistPatch, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    cl = db.query(Checklist).filter(Checklist.id == cl_id).first()
    if not cl: raise HTTPException(404, "Checklist introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(cl, k, v)
    db.commit(); db.refresh(cl)
    return cl


@router.delete("/checklists/{cl_id}", status_code=204)
def delete_checklist(cl_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    cl = db.query(Checklist).filter(Checklist.id == cl_id).first()
    if not cl: raise HTTPException(404, "Checklist introuvable")
    db.delete(cl); db.commit()


@router.post("/checklists/{cl_id}/items", status_code=201)
def add_item(cl_id: int, data: ChecklistItemIn, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    if not db.query(Checklist).filter(Checklist.id == cl_id).first():
        raise HTTPException(404, "Checklist introuvable")
    item = ChecklistItem(checklist_id=cl_id, titre=data.titre, ordre=data.ordre)
    db.add(item); db.commit(); db.refresh(item)
    return item


@router.patch("/checklists/{cl_id}/items/{item_id}")
def update_item(cl_id: int, item_id: int, data: ItemPatch, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    item = db.query(ChecklistItem).filter(ChecklistItem.id == item_id, ChecklistItem.checklist_id == cl_id).first()
    if not item: raise HTTPException(404, "Item introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    db.commit(); db.refresh(item)
    return item


@router.delete("/checklists/{cl_id}/items/{item_id}", status_code=204)
def delete_item(cl_id: int, item_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    item = db.query(ChecklistItem).filter(ChecklistItem.id == item_id, ChecklistItem.checklist_id == cl_id).first()
    if not item: raise HTTPException(404, "Item introuvable")
    db.delete(item); db.commit()


# ── Export Planning ───────────────────────────────────────────────────────────

@router.get("/export")
def export_planning(
    date_debut:  Optional[date] = Query(None),
    date_fin:    Optional[date] = Query(None),
    statut:      Optional[str]  = Query(None),
    responsable: Optional[str]  = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    BLUE   = "1F3864"
    BLUE2  = "2C4A7C"
    WHITE  = "FFFFFF"
    GREEN  = "D9EAD3"
    GRAY   = "F5F5F5"

    hdr_fill  = PatternFill("solid", fgColor=BLUE)
    hdr_font  = Font(bold=True, color=WHITE, name="Arial", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center")
    grp_fill  = PatternFill("solid", fgColor=BLUE2)
    grp_font  = Font(bold=True, color=WHITE, name="Arial", size=10)
    ok_fill   = PatternFill("solid", fgColor=GREEN)
    alt_fill  = PatternFill("solid", fgColor=GRAY)
    thin = Side(style="thin", color="CCCCCC")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    # ── Feuille 1 : Tâches ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tâches"
    headers1 = ["Responsable", "Tâche", "Date planifiée", "Date de fin", "Priorité", "Statut"]
    widths1  = [18, 40, 15, 15, 12, 14]
    for col, (h, w) in enumerate(zip(headers1, widths1), 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = brd
        ws1.column_dimensions[c.column_letter].width = w
    ws1.row_dimensions[1].height = 22

    q = db.query(Tache)
    if date_debut:  q = q.filter(Tache.date_planifiee >= date_debut)
    if date_fin:    q = q.filter(Tache.date_planifiee <= date_fin)
    if statut:      q = q.filter(Tache.statut == statut)
    if responsable: q = q.filter(Tache.responsable.ilike(f"%{responsable}%"))
    taches = q.order_by(Tache.date_planifiee.asc()).all()

    statut_labels = {"A_FAIRE": "À faire", "EN_COURS": "En cours", "TERMINE": "Terminé", "ANNULE": "Annulé"}
    for r, t in enumerate(taches, 2):
        row = [
            t.responsable or "",
            t.titre,
            t.date_planifiee.isoformat() if t.date_planifiee else "",
            t.date_fin.isoformat() if t.date_fin else "",
            t.priorite.value if hasattr(t.priorite, "value") else str(t.priorite),
            statut_labels.get(str(t.statut).split(".")[-1], str(t.statut)),
        ]
        fill = ok_fill if str(t.statut).endswith("TERMINE") else (alt_fill if r % 2 == 0 else None)
        for col, val in enumerate(row, 1):
            c = ws1.cell(row=r, column=col, value=val)
            c.border = brd
            if fill: c.fill = fill

    # ── Feuille 2 : Checklists ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Checklists")
    headers2 = ["Checklist", "Responsable", "Avancement", "Item", "Statut item"]
    widths2  = [35, 18, 14, 40, 14]
    for col, (h, w) in enumerate(zip(headers2, widths2), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = brd
        ws2.column_dimensions[c.column_letter].width = w
    ws2.row_dimensions[1].height = 22

    from sqlalchemy.orm import joinedload as jl
    checklists = db.query(Checklist).options(jl(Checklist.items)).order_by(Checklist.created_at.desc()).all()
    r2 = 2
    for cl in checklists:
        total  = len(cl.items)
        coches = sum(1 for it in cl.items if it.cochee)
        pct    = f"{round(coches/total*100)}%" if total else "—"
        done   = total > 0 and coches == total
        n = max(total, 1)
        for i, item in enumerate(cl.items if cl.items else [None]):
            row = [
                cl.nom if i == 0 else "",
                (cl.responsable or "") if i == 0 else "",
                f"{coches}/{total} ({pct})" if i == 0 else "",
                item.titre if item else "",
                ("✓ Fait" if item.cochee else "À faire") if item else "",
            ]
            fill = ok_fill if (item and item.cochee) else (grp_fill if i == 0 and done else None)
            for col, val in enumerate(row, 1):
                c = ws2.cell(row=r2, column=col, value=val)
                c.border = brd
                if i == 0 and col <= 3:
                    c.font = grp_font if done else Font(bold=True, name="Arial", size=10)
                if fill: c.fill = fill
            r2 += 1
        if n > 1:
            for col in range(1, 4):
                ws2.merge_cells(start_row=r2-n, end_row=r2-1, start_column=col, end_column=col)
                ws2.cell(row=r2-n, column=col).alignment = Alignment(vertical="center", wrap_text=True)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=planning.xlsx"},
    )
