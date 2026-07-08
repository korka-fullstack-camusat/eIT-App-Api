from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import Optional
from datetime import date as dt_date
from ..database import get_db
from ..models.materiel import Materiel, StatutMateriel, TypeMateriel, EtatMateriel
from ..models.attribution import Attribution, StatutAttribution
from ..schemas.materiel import MaterielCreate, MaterielUpdate, MaterielOut, AttributionActiveInfo
from ..models.user import User
from ..services.auth_service import require_editor

router = APIRouter(prefix="/api/materiels", tags=["Matériels"])


@router.get("/", response_model=list[MaterielOut])
def list_materiels(
    statut:        Optional[StatutMateriel] = None,
    type_materiel: Optional[TypeMateriel]   = None,
    etat:          Optional[EtatMateriel]   = None,
    projet:        Optional[str]            = None,
    assigne:       Optional[str]            = None,  # "OUI" | "NON"
    search:        Optional[str]            = None,
    db: Session = Depends(get_db),
):
    q = db.query(Materiel).options(joinedload(Materiel.attributions))
    if statut:
        q = q.filter(Materiel.statut == statut)
    if type_materiel:
        q = q.filter(Materiel.type_materiel == type_materiel)
    if etat:
        q = q.filter(Materiel.etat == etat)
    if projet:
        q = q.filter(Materiel.projet == projet)
    if search:
        term = f"%{search}%"
        q = q.outerjoin(
            Attribution,
            (Attribution.materiel_id == Materiel.id) & (Attribution.statut == StatutAttribution.ACTIVE),
        ).filter(
            Materiel.marque.ilike(term) |
            Materiel.modele.ilike(term) |
            Materiel.numero_serie.ilike(term) |
            Materiel.adresse_mac.ilike(term) |
            Materiel.reference.ilike(term) |
            Materiel.numero_bon_cmd.ilike(term) |
            Materiel.beneficiaire_nom.ilike(term) |
            Materiel.beneficiaire_prenom.ilike(term) |
            Attribution.employee_nom.ilike(term) |
            Attribution.employee_prenom.ilike(term)
        ).distinct()
    materiels = q.order_by(Materiel.marque, Materiel.modele).all()

    result = []
    for m in materiels:
        item = MaterielOut.model_validate(m)
        active = next(
            (a for a in m.attributions if a.statut == StatutAttribution.ACTIVE),
            None,
        )
        if active:
            item.attribution_active = AttributionActiveInfo.model_validate(active)

        if assigne in ("OUI", "NON"):
            is_assigned = bool(item.attribution_active or item.beneficiaire_nom or item.beneficiaire_prenom)
            if (assigne == "OUI") != is_assigned:
                continue

        result.append(item)
    return result


@router.post("/", response_model=MaterielOut, status_code=201)
def create_materiel(data: MaterielCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = Materiel(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


# NOTE : toutes les routes statiques (export-excel, stats, import…) sont déclarées
# AVANT /{materiel_id} pour éviter que FastAPI ne les capture comme un ID entier.

@router.get("/export-excel")
def export_materiels_excel_route(
    statut:        Optional[StatutMateriel] = Query(None),
    type_materiel: Optional[TypeMateriel]   = Query(None),
    etat:          Optional[EtatMateriel]   = Query(None),
    search:        Optional[str]            = Query(None),
    date_debut:    Optional[dt_date]        = Query(None),
    date_fin:      Optional[dt_date]        = Query(None),
    cols:          Optional[str]            = Query(None),
    db: Session = Depends(get_db),
):
    """Proxy vers la logique d'export — défini ici pour précéder /{materiel_id}."""
    return export_materiels_excel(
        statut=statut, type_materiel=type_materiel, etat=etat,
        search=search, date_debut=date_debut, date_fin=date_fin,
        cols=cols, db=db,
    )


@router.get("/{materiel_id}", response_model=MaterielOut)
def get_materiel(materiel_id: int, db: Session = Depends(get_db)):
    obj = db.query(Materiel).filter(Materiel.id == materiel_id).first()
    if not obj:
        raise HTTPException(404, "Matériel introuvable")
    return obj


@router.patch("/{materiel_id}", response_model=MaterielOut)
def update_materiel(materiel_id: int, data: MaterielUpdate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Materiel).filter(Materiel.id == materiel_id).first()
    if not obj:
        raise HTTPException(404, "Matériel introuvable")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(obj, k, v)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/{materiel_id}", status_code=204)
def delete_materiel(materiel_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Materiel).filter(Materiel.id == materiel_id).first()
    if not obj:
        raise HTTPException(404, "Matériel introuvable")
    if obj.statut == StatutMateriel.ATTRIBUE:
        raise HTTPException(400, "Impossible de supprimer un matériel attribué")
    db.delete(obj)
    db.commit()


@router.post("/import")
def import_materiels(file: UploadFile = File(...), db: Session = Depends(get_db), _: User = Depends(require_editor)):
    content = file.file.read()
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        return _import_materiels_xlsx(content, db)
    return _import_materiels_csv(content, db)


def _import_materiels_csv(content: bytes, db: Session):
    import csv, io
    from ..models.materiel import TypeMateriel, EtatMateriel

    TYPE_MAP = {
        "PC PORTABLE": "ORDINATEUR_PORTABLE", "PC FIXE": "ORDINATEUR_FIXE",
        "ORDINATEUR PORTABLE": "ORDINATEUR_PORTABLE", "ORDINATEUR FIXE": "ORDINATEUR_FIXE",
        "ORDINATEUR_PORTABLE": "ORDINATEUR_PORTABLE", "ORDINATEUR_FIXE": "ORDINATEUR_FIXE",
        "ECRAN": "ECRAN", "ÉCRAN": "ECRAN", "SOURIS": "SOURIS", "CLAVIER": "CLAVIER",
        "TELEPHONE": "TELEPHONE", "TÉLÉPHONE": "TELEPHONE", "IMPRIMANTE": "IMPRIMANTE",
        "SWITCH": "SWITCH", "ROUTEUR": "ROUTEUR", "ONDULEUR": "ONDULEUR", "AUTRE": "AUTRE",
    }
    ETAT_MAP = {
        "NEUF": "NEUF", "BON": "BON", "USAGE": "USAGE", "USAGÉ": "USAGE",
        "DEFECTUEUX": "DEFECTUEUX", "DÉFECTUEUX": "DEFECTUEUX",
    }

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    # Normalize header keys
    def norm(s: str) -> str:
        return s.strip().upper().replace("°", "").replace(" ", "_").replace("É", "E").replace("È", "E")

    created, errors = 0, []
    for i, raw_row in enumerate(reader, start=2):
        row = {norm(k): (v.strip() if v else "") for k, v in raw_row.items()}
        type_raw  = row.get("TYPE", "")
        type_val  = TYPE_MAP.get(type_raw.upper(), None)
        if not type_val:
            errors.append({"ligne": i, "message": f"Type inconnu : '{type_raw}'"})
            continue
        marque = row.get("MARQUE", "").strip()
        if not marque:
            errors.append({"ligne": i, "message": "Marque manquante"})
            continue
        etat_raw = row.get("ETAT", row.get("ÉTAT", "BON")).upper()
        etat_val = ETAT_MAP.get(etat_raw, "BON")

        acq_raw = row.get("ACQUISITION", row.get("DATE", "")).strip()
        acq = None
        if acq_raw:
            from datetime import date as dt_date, datetime as dt_datetime
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    acq = dt_datetime.strptime(acq_raw, fmt).date()
                    break
                except ValueError:
                    pass

        obj = Materiel(
            type_materiel    = TypeMateriel(type_val),
            marque           = marque,
            modele           = row.get("MODELE", row.get("MODÈLE", "")) or None,
            numero_serie     = row.get("N_SERIE", row.get("N__SERIE", "")) or None,
            adresse_mac      = row.get("ADRESSE_MAC", row.get("ADRESSE_IP", "")) or None,
            numero_bon_cmd   = row.get("N_PO", row.get("N__PO", "")) or None,
            etat             = EtatMateriel(etat_val),
            date_acquisition = acq,
        )
        db.add(obj)
        created += 1

    db.commit()
    return {"created": created, "errors": errors, "total_lignes": created + len(errors)}


def _import_materiels_xlsx(content: bytes, db: Session):
    """Import depuis un fichier 'Suivi Parc' (.xlsx) — ne récupère que les
    informations relatives au matériel (nature, désignation, n° série,
    adresse MAC, n° PO, projet, date d'attribution, statut)."""
    import io, re, unicodedata
    from datetime import datetime as dt_datetime, date as dt_date2
    from openpyxl import load_workbook
    from ..models.materiel import TypeMateriel, EtatMateriel, StatutMateriel

    NATURE_MAP = {
        "PC": "ORDINATEUR_PORTABLE",
        "ORDINATEUR": "ORDINATEUR_PORTABLE",
        "ORDINATEURPORTABLE": "ORDINATEUR_PORTABLE",
        "ORDINATEURFIXE": "ORDINATEUR_FIXE",
        "PCPORTABLE": "ORDINATEUR_PORTABLE",
        "PCFIXE": "ORDINATEUR_FIXE",
        "TELEPHONE": "TELEPHONE",
        "TABLETTE": "TABLETTE",
        "ECRAN": "ECRAN",
        "SOURIS": "SOURIS",
        "CLAVIER": "CLAVIER",
        "CASQUE": "AUTRE",
        "IMPRIMANTE": "IMPRIMANTE",
        "SWITCH": "SWITCH",
        "ROUTEUR": "ROUTEUR",
        "ONDULEUR": "ONDULEUR",
        "AP": "AP",
        "POINTDACCES": "AP",
        "SERVEUR": "SERVEUR",
        "PAREFEU": "PARE_FEU",
        "FIREWALL": "PARE_FEU",
    }
    STATUT_MAP = {
        "ENSERVICE": "ATTRIBUE",
        "ENPANNE": "EN_PANNE",
        "ENSTOCK": "DISPONIBLE",
        "REFORME": "REFORME",
        "MAINTENANCE": "MAINTENANCE",
        "ENMAINTENANCE": "MAINTENANCE",
        "RESTITUE": "DISPONIBLE",
    }

    def norm(s) -> str:
        s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
        return re.sub(r"[^A-Z0-9]", "", s.upper())

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Fichier Excel illisible.")

    sheet_name = "Suivi Parc Informatique" if "Suivi Parc Informatique" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Repérer la ligne d'en-tête (celle qui contient "Nature", "Désignation" ou "Matricule")
    HEADER_ANCHORS = {"NATURE", "DESIGNATIONEQUIPEMENT", "DESIGNATION", "MATRICULE"}
    header_row_idx, headers = None, {}
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        normalized = [norm(v) if v is not None else "" for v in row]
        if HEADER_ANCHORS & set(normalized):
            header_row_idx = r_idx
            headers = {h: c for c, h in enumerate(normalized) if h}
            break
    if header_row_idx is None:
        raise HTTPException(400, "Format non reconnu : aucune colonne attendue (Nature, Désignation, Matricule) trouvée dans les 10 premières lignes.")

    def col(row, *names):
        for name in names:
            idx = headers.get(name)
            if idx is not None and idx < len(row):
                v = row[idx]
                if isinstance(v, str):
                    v = v.strip()
                if v not in (None, ""):
                    return v
        return None

    existing_series = {
        s for (s,) in db.query(Materiel.numero_serie).filter(Materiel.numero_serie.isnot(None)).all()
    }

    created, errors = 0, []
    for i, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        if all(v is None for v in row):
            continue

        nature = col(row, "NATURE")
        if not nature:
            continue

        type_val = NATURE_MAP.get(norm(nature))
        if not type_val:
            errors.append({"ligne": i, "message": f"Nature non reconnue : '{nature}'"})
            continue

        designation = col(row, "DESIGNATIONEQUIPEMENT") or ""
        designation = str(designation).strip()
        if designation:
            parts = designation.split(maxsplit=1)
            marque, modele = parts[0], (parts[1] if len(parts) > 1 else "")
        else:
            marque, modele = TYPE_LABELS_FALLBACK.get(type_val, "Matériel"), ""

        numero_serie = col(row, "NSERIE", "NDESERIE")
        numero_serie = str(numero_serie).strip() if numero_serie else None
        if numero_serie and numero_serie in existing_series:
            errors.append({"ligne": i, "message": f"N° série déjà existant : '{numero_serie}'"})
            continue

        adresse_mac    = col(row, "REFCARTERESEAU")
        numero_bon_cmd = col(row, "PO")
        projet         = col(row, "PROJET")
        matricule      = col(row, "MATRICULE")
        nom            = col(row, "NOM")
        prenom         = col(row, "PRENOM")

        date_attr_raw = col(row, "DATEDATTRIBUTION", "DATE")
        date_acquisition = None
        if isinstance(date_attr_raw, (dt_datetime, dt_date2)):
            date_acquisition = date_attr_raw.date() if isinstance(date_attr_raw, dt_datetime) else date_attr_raw
        elif isinstance(date_attr_raw, str) and date_attr_raw.strip():
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    date_acquisition = dt_datetime.strptime(date_attr_raw.strip(), fmt).date()
                    break
                except ValueError:
                    pass

        statut_raw = col(row, "STATUT")
        statut_val = STATUT_MAP.get(norm(statut_raw), "DISPONIBLE") if statut_raw else "DISPONIBLE"

        obj = Materiel(
            type_materiel    = TypeMateriel(type_val),
            marque           = str(marque)[:100],
            modele           = str(modele)[:150] or None,
            numero_serie     = numero_serie,
            adresse_mac      = str(adresse_mac)[:50] if adresse_mac else None,
            numero_bon_cmd   = str(numero_bon_cmd)[:100] if numero_bon_cmd else None,
            projet           = str(projet)[:100] if projet else None,
            beneficiaire_matricule = str(matricule)[:50] if matricule else None,
            beneficiaire_nom       = str(nom)[:100] if nom else None,
            beneficiaire_prenom    = str(prenom)[:100] if prenom else None,
            etat             = EtatMateriel.BON,
            statut           = StatutMateriel(statut_val),
            date_acquisition = date_acquisition,
        )
        db.add(obj)
        if numero_serie:
            existing_series.add(numero_serie)
        created += 1

    db.commit()
    return {"created": created, "errors": errors, "total_lignes": created + len(errors)}


TYPE_LABELS_FALLBACK = {
    "ORDINATEUR_PORTABLE": "PC", "ORDINATEUR_FIXE": "PC", "ECRAN": "Écran",
    "SOURIS": "Souris", "CLAVIER": "Clavier", "TELEPHONE": "Téléphone",
    "TABLETTE": "Tablette", "IMPRIMANTE": "Imprimante", "SWITCH": "Switch",
    "ROUTEUR": "Routeur", "ONDULEUR": "Onduleur", "AUTRE": "Matériel",
}


def export_materiels_excel(
    statut:        Optional[StatutMateriel],
    type_materiel: Optional[TypeMateriel],
    etat:          Optional[EtatMateriel],
    search:        Optional[str],
    date_debut:    Optional[dt_date],
    date_fin:      Optional[dt_date],
    cols:          Optional[str],
    db,
):
    """Génère un fichier Excel stylisé avec en-têtes bleus."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter

    # ── Requête ──────────────────────────────────────────────────────────────────
    q = db.query(Materiel).options(joinedload(Materiel.attributions))
    if statut:        q = q.filter(Materiel.statut == statut)
    if type_materiel: q = q.filter(Materiel.type_materiel == type_materiel)
    if etat:          q = q.filter(Materiel.etat == etat)
    if search:
        term = f"%{search}%"
        q = q.outerjoin(
            Attribution,
            (Attribution.materiel_id == Materiel.id) & (Attribution.statut == StatutAttribution.ACTIVE),
        ).filter(
            Materiel.marque.ilike(term) | Materiel.modele.ilike(term) |
            Materiel.numero_serie.ilike(term) | Materiel.adresse_mac.ilike(term) |
            Materiel.reference.ilike(term) |
            Materiel.beneficiaire_nom.ilike(term) | Materiel.beneficiaire_prenom.ilike(term) |
            Attribution.employee_nom.ilike(term) | Attribution.employee_prenom.ilike(term)
        ).distinct()
    materiels = q.order_by(Materiel.marque).all()

    # Filtre date côté Python (date_acquisition peut être null)
    if date_debut:
        materiels = [m for m in materiels if m.date_acquisition and m.date_acquisition >= date_debut]
    if date_fin:
        materiels = [m for m in materiels if m.date_acquisition and m.date_acquisition <= date_fin]

    # Attribution active par matériel
    def get_active(m: Materiel):
        return next((a for a in m.attributions if a.statut == StatutAttribution.ACTIVE), None)

    TYPE_LABELS = {
        "ORDINATEUR_PORTABLE": "PC Portable", "ORDINATEUR_FIXE": "PC Fixe",
        "ECRAN": "Écran", "SOURIS": "Souris", "CLAVIER": "Clavier",
        "TELEPHONE": "Téléphone", "IMPRIMANTE": "Imprimante",
        "SWITCH": "Switch", "ROUTEUR": "Routeur", "ONDULEUR": "Onduleur", "AUTRE": "Autre",
    }
    STATUT_LABELS = {
        "DISPONIBLE": "Disponible", "ATTRIBUE": "Attribué",
        "MAINTENANCE": "Maintenance", "EN_PANNE": "En panne", "REFORME": "Réformé",
    }

    # ── Définition des colonnes ───────────────────────────────────────────────
    ALL_COLS = [
        ("id",          "ID",               lambda m, a: m.id),
        ("type",        "Type",             lambda m, a: TYPE_LABELS.get(m.type_materiel, m.type_materiel)),
        ("marque",      "Marque",           lambda m, a: m.marque or ""),
        ("modele",      "Modèle",           lambda m, a: m.modele or ""),
        ("serie",       "N° Série",         lambda m, a: m.numero_serie or ""),
        ("mac",         "Adresse MAC",      lambda m, a: m.adresse_mac or ""),
        ("reference",   "Référence",        lambda m, a: m.reference or ""),
        ("po",          "N° PO",            lambda m, a: m.numero_bon_cmd or ""),
        ("etat",        "État",             lambda m, a: m.etat or ""),
        ("statut",      "Statut",           lambda m, a: STATUT_LABELS.get(m.statut, m.statut or "")),
        ("acquisition", "Date Acquisition", lambda m, a: m.date_acquisition.strftime("%d/%m/%Y") if m.date_acquisition else ""),
        ("assigne",     "Assigné à",        lambda m, a: f"{a.employee_prenom or ''} {a.employee_nom}".strip() if a else ""),
    ]

    selected_keys = set(cols.split(",")) if cols else {c[0] for c in ALL_COLS}
    columns = [(k, lbl, fn) for k, lbl, fn in ALL_COLS if k in selected_keys]

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Parc Informatique"

    # Couleurs
    BLUE_HDR   = "1B3D6F"   # bleu foncé camublue
    WHITE      = "FFFFFF"
    ROW_EVEN   = "EEF4FF"   # bleu très clair
    ROW_ODD    = "FFFFFF"
    BORDER_COL = "C5D3E8"   # bleu gris léger

    thin = Side(style="thin", color=BORDER_COL)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Ligne de titre ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value  = "INVENTAIRE DU PARC INFORMATIQUE"
    title_cell.font   = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title_cell.fill   = PatternFill("solid", fgColor=BLUE_HDR)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Ligne sous-titre (date + filtres) ─────────────────────────────────────
    from datetime import datetime
    subtitle_parts = [f"Exporté le {datetime.now().strftime('%d/%m/%Y')}"]
    if date_debut: subtitle_parts.append(f"Du {date_debut.strftime('%d/%m/%Y')}")
    if date_fin:   subtitle_parts.append(f"au {date_fin.strftime('%d/%m/%Y')}")
    if statut:     subtitle_parts.append(f"Statut : {STATUT_LABELS.get(statut, statut)}")
    if type_materiel: subtitle_parts.append(f"Type : {TYPE_LABELS.get(type_materiel, type_materiel)}")
    subtitle_parts.append(f"{len(materiels)} matériel(s)")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sub_cell = ws.cell(row=2, column=1)
    sub_cell.value = "  |  ".join(subtitle_parts)
    sub_cell.font  = Font(name="Calibri", italic=True, size=9, color="5B7DB1")
    sub_cell.fill  = PatternFill("solid", fgColor="D9E6F7")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # ── Ligne vide séparatrice ────────────────────────────────────────────────
    ws.row_dimensions[3].height = 6

    # ── En-têtes de colonnes ──────────────────────────────────────────────────
    HDR_ROW = 4
    for ci, (_, lbl, _fn) in enumerate(columns, start=1):
        c = ws.cell(row=HDR_ROW, column=ci, value=lbl)
        c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        c.border    = border_all
    ws.row_dimensions[HDR_ROW].height = 24

    # Figer la ligne d'en-tête
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # ── Données ───────────────────────────────────────────────────────────────
    for ri, m in enumerate(materiels):
        row_num  = HDR_ROW + 1 + ri
        fill_color = ROW_EVEN if ri % 2 == 0 else ROW_ODD
        row_fill   = PatternFill("solid", fgColor=fill_color)
        active_attr = get_active(m)
        for ci, (_, _lbl, fn) in enumerate(columns, start=1):
            val = fn(m, active_attr)
            c = ws.cell(row=row_num, column=ci, value=val)
            c.fill      = row_fill
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            c.border    = border_all
        ws.row_dimensions[row_num].height = 18

    # ── Largeurs automatiques ─────────────────────────────────────────────────
    for ci, (_, lbl, _) in enumerate(columns, start=1):
        col_letter = get_column_letter(ci)
        # Largeur basée sur la largeur du titre + données
        all_vals = [lbl] + [
            str(fn(m, get_active(m))) for m in materiels
        ]
        _, _lbl2, fn2 = columns[ci - 1]
        max_len = max((len(str(fn2(m, get_active(m)))) for m in materiels), default=0)
        width   = max(len(lbl) + 2, min(max_len + 2, 40))
        ws.column_dimensions[col_letter].width = width

    # ── Pied de page ──────────────────────────────────────────────────────────
    ws.oddFooter.center.text  = f"&\"Calibri\"&8 CAMUSAT — Parc Informatique  |  Page &P / &N"
    ws.oddFooter.center.size  = 8

    # ── Sauvegarde ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = "materiels"
    if date_debut: fname += f"_du_{date_debut}"
    if date_fin:   fname += f"_au_{date_fin}"
    fname += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/stats/summary")
def stats_summary(db: Session = Depends(get_db)):
    total       = db.query(Materiel).count()
    disponible  = db.query(Materiel).filter(Materiel.statut == StatutMateriel.DISPONIBLE).count()
    attribue    = db.query(Materiel).filter(Materiel.statut == StatutMateriel.ATTRIBUE).count()
    maintenance = db.query(Materiel).filter(Materiel.statut == StatutMateriel.MAINTENANCE).count()
    en_panne    = db.query(Materiel).filter(Materiel.statut == StatutMateriel.EN_PANNE).count()
    reforme     = db.query(Materiel).filter(Materiel.statut == StatutMateriel.REFORME).count()
    return {
        "total": total, "disponible": disponible, "attribue": attribue,
        "maintenance": maintenance, "en_panne": en_panne, "reforme": reforme,
    }


@router.get("/stats/par-type")
def stats_par_type(db: Session = Depends(get_db)):
    rows = (
        db.query(Materiel.type_materiel, func.count(Materiel.id))
        .group_by(Materiel.type_materiel)
        .order_by(func.count(Materiel.id).desc())
        .all()
    )
    return [{"type": r[0], "count": r[1]} for r in rows]


@router.get("/stats/par-marque")
def stats_par_marque(db: Session = Depends(get_db)):
    rows = (
        db.query(Materiel.marque, func.count(Materiel.id))
        .group_by(Materiel.marque)
        .order_by(func.count(Materiel.id).desc())
        .limit(8)
        .all()
    )
    return [{"marque": r[0], "count": r[1]} for r in rows]


@router.get("/stats/par-projet")
def stats_par_projet(db: Session = Depends(get_db)):
    rows = (
        db.query(Materiel.projet, func.count(Materiel.id))
        .filter(Materiel.projet.isnot(None), Materiel.projet != "")
        .group_by(Materiel.projet)
        .order_by(Materiel.projet)
        .all()
    )
    return [{"projet": r[0], "count": r[1]} for r in rows]
