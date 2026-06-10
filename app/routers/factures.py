from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional
from decimal import Decimal, InvalidOperation
import re
import io
import unicodedata
import pandas as pd
from ..database import get_db
from ..models.facture import FactureTelecom, LigneFacture
from ..models.telephonie import NumeroSIM
from ..schemas.facture import FactureOut, ImportResult
from ..models.user import User
from ..services.auth_service import require_editor

router = APIRouter(prefix="/api/factures", tags=["Factures télécom"])


# ── Helpers récapitulatif facture ──────────────────────────────────────────────

def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def _norm_header(val) -> str:
    """Normalise un en-tête de colonne : majuscule, sans accent, sans saut de ligne, espaces réduits."""
    s = str(val or "").replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip().upper()
    return _strip_accents(s)


def _parse_number(val) -> Optional[Decimal]:
    """Parse un nombre type '9 800', '- 24', '1 493,00' → Decimal."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() == "nan":
        return None
    s = s.replace(" ", "").replace(" ", "").replace(",", ".")
    # Gère le cas "-24" ou "- 24" (espace déjà retiré)
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


# Mapping des en-têtes (forme normalisée) → champ du modèle.
# L'ordre compte : les libellés les plus spécifiques sont testés en premier.
LINE_FIELD_MATCHERS = [
    ("reference_facture", lambda h: "REFERENCE" in h and "FACTURE" in h),
    ("solde_facture",     lambda h: "SOLDE" in h and "FACTURE" in h),
    ("montant_ht_rutel",  lambda h: "RUTEL" in h and "TVA" in h),
    ("rutel",             lambda h: "RUTEL" in h),
    ("montant_ht",        lambda h: "MONTANT" in h and ("HORS TAXE" in h or "HT" in h)),
    ("montant_ttc",       lambda h: "MONTANT" in h and "TTC" in h),
    ("tva",               lambda h: "TVA" in h),
    ("arrondi_precedent", lambda h: "ARRONDI" in h and "PRECEDENT" in h),
    ("arrondi_encours",   lambda h: "ARRONDI" in h and ("EN COURS" in h or "ENCOURS" in h)),
    ("numero",            lambda h: h == "NUMERO" or h.startswith("NUMERO ") or h.startswith("N°") or "TELEPHONE" in h),
    ("type_ligne",        lambda h: h == "TYPE" or "TYPE" in h),
]

STRING_FIELDS = ["reference_facture", "type_ligne"]

NUMERIC_FIELDS = ["montant_ht", "rutel", "montant_ht_rutel", "tva", "montant_ttc",
                  "arrondi_precedent", "arrondi_encours", "solde_facture"]


def _extract_lignes_table(content: bytes, filename: str) -> tuple[list[dict], dict]:
    """
    Cherche dans le fichier la ligne d'en-têtes du récapitulatif facture
    (Numéro, Référence Facture, Montant Hors Taxe, Rutel (5%), Hors TVA avec
    Rutel, TVA (18%), Montant TTC, Arrondi Précédent, Arrondi En cours,
    Solde Facture) et renvoie :
      - TOUTES les lignes de données qui suivent (une par numéro), avec leurs
        valeurs telles que dans le fichier ;
      - les totaux de la facture, extraits tels quels des lignes "Totaux" et
        "Total facture (FCFA)" du pied de tableau.
    """
    try:
        if filename.endswith(".csv"):
            raw = pd.read_csv(io.BytesIO(content), sep=";", header=None, dtype=str)
        else:
            raw = pd.read_excel(io.BytesIO(content), header=None, dtype=str)
    except Exception:
        return [], {}

    n_rows, n_cols = raw.shape
    for ri in range(n_rows):
        row_vals = [_norm_header(raw.iat[ri, ci]) for ci in range(n_cols)]
        col_map = {}
        for ci, h in enumerate(row_vals):
            if not h:
                continue
            for field, matcher in LINE_FIELD_MATCHERS:
                if field not in col_map and matcher(h):
                    col_map[field] = ci
                    break
        # On considère que c'est la ligne d'en-têtes si on reconnaît au moins
        # "Numéro" + "Montant TTC" (les 2 colonnes-clés).
        if "numero" in col_map and "montant_ttc" in col_map:
            lines = []
            totals: dict = {}
            for di in range(ri + 1, n_rows):
                row_label = ""
                for ci in range(n_cols):
                    cell_label = _norm_header(raw.iat[di, ci])
                    if "TOTAUX" in cell_label or "TOTAL FACTURE" in cell_label:
                        row_label = cell_label
                        break

                if row_label:
                    # "Total facture (FCFA)" peut se répéter (une fois par page
                    # imprimée) avec la même valeur : on ne garde que la première
                    # occurrence comme solde de la facture. Les lignes "Totaux"
                    # ne sont que des sous-totaux de page et ne sont pas utilisées.
                    if "TOTAL FACTURE" in row_label and "solde_facture" not in totals:
                        for field, ci in col_map.items():
                            if field in STRING_FIELDS or field == "numero":
                                continue
                            val = _parse_number(raw.iat[di, ci])
                            if val is not None:
                                totals["solde_facture"] = val
                                break
                    continue

                num_val = raw.iat[di, col_map["numero"]]
                if num_val is None or str(num_val).strip() == "" or str(num_val).strip().lower() == "nan":
                    continue
                numero_raw = str(num_val).strip()

                digits = numero_raw.replace(" ", "").replace(" ", "")
                if not digits.isdigit() or not (6 <= len(digits) <= 15):
                    continue
                line = {"numero_raw": numero_raw}
                for field, ci in col_map.items():
                    if field == "numero":
                        continue
                    val = raw.iat[di, ci]
                    if field in STRING_FIELDS:
                        sval = str(val).strip() if val is not None else ""
                        line[field] = sval if sval and sval.lower() != "nan" else None
                    else:
                        line[field] = _parse_number(val)
                lines.append(line)
            if lines:
                return lines, totals
    return [], {}


def _extract_compte_client(content: bytes, filename: str) -> Optional[str]:
    """Cherche une ligne 'N° compte client : XXXXXXXX' dans l'en-tête du fichier."""
    try:
        if filename.endswith(".csv"):
            raw = pd.read_csv(io.BytesIO(content), sep=";", header=None, dtype=str)
        else:
            raw = pd.read_excel(io.BytesIO(content), header=None, dtype=str)
    except Exception:
        return None

    n_rows, n_cols = raw.shape
    for ri in range(min(25, n_rows)):
        for ci in range(n_cols - 1):
            label = _norm_header(raw.iat[ri, ci])
            if "COMPTE" in label and "CLIENT" in label:
                val = raw.iat[ri, ci + 1]
                if val is not None and str(val).strip() and str(val).strip().lower() != "nan":
                    return str(val).strip()
    return None


def _attach_ecart(factures: list[FactureTelecom]) -> list[FactureOut]:
    """Calcule l'écart de Montant TTC d'une facture par rapport au mois précédent
    (chronologiquement, tous opérateurs confondus)."""
    ordered = sorted(factures, key=lambda f: (f.annee, f.mois))
    prev_montant: dict[int, Decimal] = {}
    out_by_id: dict[int, FactureOut] = {}
    for f in ordered:
        out = FactureOut.model_validate(f)
        idx = f.annee * 12 + f.mois
        prev_key = idx - 1
        if prev_key in prev_montant and out.montant_ttc is not None:
            prev = prev_montant[prev_key]
            out.ecart = out.montant_ttc - prev
            if prev != 0:
                out.ecart_pct = float(out.ecart / prev * 100)
        if out.montant_ttc is not None:
            prev_montant[idx] = out.montant_ttc
        out_by_id[f.id] = out
    return out_by_id


@router.get("/", response_model=list[FactureOut])
def list_factures(annee: Optional[int] = None, db: Session = Depends(get_db)):
    # On charge toutes les factures pour calculer l'écart par rapport au mois
    # précédent même s'il appartient à une autre année, puis on filtre.
    all_factures = db.query(FactureTelecom).order_by(FactureTelecom.annee, FactureTelecom.mois).all()
    out_by_id = _attach_ecart(all_factures)
    selected = [f for f in all_factures if not annee or f.annee == annee]
    selected.sort(key=lambda f: (f.annee, f.mois), reverse=True)
    return [out_by_id[f.id] for f in selected]


@router.get("/export-excel")
def export_factures_excel(
    annee: Optional[int] = Query(None),
    mois:  Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """Export Excel stylisé — une ligne par ligne de facture."""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    MOIS_LABELS = ["","Janvier","Février","Mars","Avril","Mai","Juin",
                   "Juillet","Août","Septembre","Octobre","Novembre","Décembre"]

    q = db.query(FactureTelecom)
    if annee: q = q.filter(FactureTelecom.annee == annee)
    if mois:  q = q.filter(FactureTelecom.mois == mois)
    factures = q.order_by(FactureTelecom.annee.desc(), FactureTelecom.mois.desc()).all()

    BLUE_HDR = "1B3D6F"; WHITE = "FFFFFF"; ROW_EVEN = "EEF4FF"; BORDER_COL = "C5D3E8"
    thin = Side(style="thin", color=BORDER_COL)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook(); ws = wb.active; ws.title = "Factures Télécom"

    HEADERS = ["Période", "Opérateur", "Fichier", "Numéro", "Montant (FCFA)", "Reconnu", "Notes"]

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    tc = ws.cell(row=1, column=1, value="FACTURES TÉLÉCOM")
    tc.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=BLUE_HDR)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Sous-titre
    parts = [f"Exporté le {datetime.now().strftime('%d/%m/%Y')}"]
    if annee: parts.append(f"Année {annee}")
    if mois:  parts.append(MOIS_LABELS[mois])
    total_lignes = sum(len(f.lignes) for f in factures)
    parts.append(f"{len(factures)} facture(s) · {total_lignes} ligne(s)")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    sc = ws.cell(row=2, column=1, value="  |  ".join(parts))
    sc.font = Font(name="Calibri", italic=True, size=9, color="5B7DB1")
    sc.fill = PatternFill("solid", fgColor="D9E6F7")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # En-têtes
    HDR_ROW = 4
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=HDR_ROW, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all
    ws.row_dimensions[HDR_ROW].height = 24
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    ri = 0
    for f in factures:
        periode = f"{MOIS_LABELS[f.mois]} {f.annee}"
        for l in f.lignes:
            row_num  = HDR_ROW + 1 + ri
            row_fill = PatternFill("solid", fgColor=(ROW_EVEN if ri % 2 == 0 else WHITE))
            vals = [
                periode,
                f.operateur or "",
                f.nom_fichier or "",
                l.numero_raw or "",
                float(l.montant) if l.montant else 0,
                "Oui" if l.sim_id else "Non",
                f.notes or "",
            ]
            for ci, v in enumerate(vals, start=1):
                c = ws.cell(row=row_num, column=ci, value=v)
                c.fill = row_fill
                c.font = Font(name="Calibri", size=10)
                c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
                c.border = border_all
                # Montant aligné à droite
                if ci == 5:
                    c.alignment = Alignment(vertical="center", horizontal="right")
                    c.number_format = "#,##0"
                # Reconnu coloré
                if ci == 6 and not l.sim_id:
                    c.font = Font(name="Calibri", size=10, color="D97706")
            ws.row_dimensions[row_num].height = 18
            ri += 1

    # Largeurs
    widths = [16, 12, 28, 18, 16, 10, 20]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.oddFooter.center.text = "&\"Calibri\"&8 CAMUSAT — Factures Télécom  |  Page &P / &N"

    # ── Feuille 2 : Récapitulatif (avec écart mois précédent) ──
    all_factures = db.query(FactureTelecom).order_by(FactureTelecom.annee, FactureTelecom.mois).all()
    out_by_id = _attach_ecart(all_factures)

    ws2 = wb.create_sheet("Récapitulatif")
    HEADERS2 = ["Période", "Numéro", "Référence Facture", "Montant Hors Taxe", "Rutel (5%)",
                "Hors TVA avec Rutel", "TVA (18%)", "Montant TTC", "Arrondi Précédent",
                "Arrondi En cours", "Solde Facture", "Écart"]
    for ci, h in enumerate(HEADERS2, start=1):
        c = ws2.cell(row=1, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border_all
    ws2.row_dimensions[1].height = 24
    ws2.freeze_panes = ws2.cell(row=2, column=1)

    ri2 = 0
    for f in factures:
        out = out_by_id.get(f.id)
        row_num = 2 + ri2
        row_fill = PatternFill("solid", fgColor=(ROW_EVEN if ri2 % 2 == 0 else WHITE))
        def num(v):
            return float(v) if v is not None else None
        ecart_str = None
        if out and out.ecart is not None:
            sign = "+" if out.ecart > 0 else ""
            ecart_str = f"{sign}{float(out.ecart):,.0f}".replace(",", " ")
            if out.ecart_pct is not None:
                ecart_str += f" ({sign}{out.ecart_pct:.1f}%)"
        vals = [
            f"{MOIS_LABELS[f.mois]} {f.annee}",
            f.numero_compte or "",
            f.reference_facture or "",
            num(f.montant_ht), num(f.rutel), num(f.montant_ht_rutel), num(f.tva),
            num(f.montant_ttc), num(f.arrondi_precedent), num(f.arrondi_encours), num(f.solde_facture),
            ecart_str or "",
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws2.cell(row=row_num, column=ci, value=v)
            c.fill = row_fill
            c.font = Font(name="Calibri", size=10)
            c.border = border_all
            if ci in (4, 5, 6, 7, 8, 9, 10, 11) and v is not None:
                c.alignment = Alignment(vertical="center", horizontal="right")
                c.number_format = "#,##0"
            else:
                c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws2.row_dimensions[row_num].height = 18
        ri2 += 1

    widths2 = [16, 14, 18, 16, 12, 18, 12, 16, 14, 14, 14, 18]
    for ci, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"factures_{annee or 'all'}"
    if mois: fname += f"_{mois:02d}"
    fname += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/{facture_id}", response_model=FactureOut)
def get_facture(facture_id: int, db: Session = Depends(get_db)):
    obj = db.query(FactureTelecom).filter(FactureTelecom.id == facture_id).first()
    if not obj:
        raise HTTPException(404, "Facture introuvable")
    all_factures = db.query(FactureTelecom).order_by(FactureTelecom.annee, FactureTelecom.mois).all()
    out_by_id = _attach_ecart(all_factures)
    return out_by_id[obj.id]


@router.post("/import", response_model=ImportResult)
async def import_facture(
    mois:      int = Form(...),
    annee:     int = Form(...),
    operateur: Optional[str] = Form(None),
    notes:     Optional[str] = Form(None),
    file:      UploadFile = File(...),
    db:        Session = Depends(get_db),
    _:         User = Depends(require_editor),
):
    # Vérifier doublon
    existing = db.query(FactureTelecom).filter(
        FactureTelecom.mois == mois, FactureTelecom.annee == annee
    ).first()
    if existing:
        raise HTTPException(400, f"Une facture existe déjà pour {mois:02d}/{annee}")

    # Lire le fichier Excel / CSV
    content = await file.read()

    # ── 1) Tableau récapitulatif détaillé (une ligne par numéro) ──
    lignes_data, file_totals = _extract_lignes_table(content, file.filename or "")
    numero_compte = _extract_compte_client(content, file.filename or "")

    # ── 2) Repli sur l'ancien format simple (colonnes Numéro / Montant) ──
    if not lignes_data:
        try:
            if file.filename.endswith(".csv"):
                df = pd.read_csv(io.BytesIO(content), sep=";", dtype=str)
            else:
                df = pd.read_excel(io.BytesIO(content), dtype=str)
            df.columns = [c.strip().upper() for c in df.columns]
            col_num  = next((c for c in df.columns if "NUM" in c or "TEL" in c), None)
            col_mont = next((c for c in df.columns if "MONT" in c or "COUT" in c or "AMOUNT" in c), None)
        except Exception:
            df, col_num, col_mont = None, None, None

        if df is not None and col_num and col_mont:
            for _, row in df.iterrows():
                numero_raw = str(row[col_num]).strip()
                montant = _parse_number(row[col_mont])
                if montant is None or not numero_raw or numero_raw.lower() == "nan":
                    continue
                lignes_data.append({"numero_raw": numero_raw, "montant_ttc": montant})

    if not lignes_data:
        raise HTTPException(400, "Format de fichier non reconnu : aucune colonne récapitulative (Numéro / Montant TTC) ni colonnes Numéro/Montant détectées.")

    # Charger tous les numéros SIM connus
    sims_map = {s.numero: s.id for s in db.query(NumeroSIM).all()}

    facture = FactureTelecom(
        mois=mois, annee=annee, operateur=operateur, notes=notes, nom_fichier=file.filename,
        numero_compte=numero_compte,
    )
    db.add(facture)
    db.flush()

    reconnus = 0
    non_reconnus_list = []
    montant_total = Decimal("0")
    sums = {field: Decimal("0") for field in NUMERIC_FIELDS}

    for ld in lignes_data:
        numero_raw = ld["numero_raw"]
        montant_ttc = ld.get("montant_ttc")
        montant = montant_ttc if montant_ttc is not None else Decimal("0")

        sim_id = sims_map.get(numero_raw)
        non_reconnu = "N" if sim_id else "O"
        if sim_id:
            reconnus += 1
        else:
            non_reconnus_list.append(numero_raw)

        db.add(LigneFacture(
            facture_id=facture.id,
            sim_id=sim_id,
            numero_raw=numero_raw,
            montant=montant,
            non_reconnu=non_reconnu,
            reference_facture=ld.get("reference_facture"),
            montant_ht=ld.get("montant_ht"),
            rutel=ld.get("rutel"),
            montant_ht_rutel=ld.get("montant_ht_rutel"),
            tva=ld.get("tva"),
            montant_ttc=ld.get("montant_ttc"),
            arrondi_precedent=ld.get("arrondi_precedent"),
            arrondi_encours=ld.get("arrondi_encours"),
            solde_facture=ld.get("solde_facture"),
            type_ligne=ld.get("type_ligne"),
        ))
        montant_total += montant
        for field in NUMERIC_FIELDS:
            v = ld.get(field)
            if v is not None:
                sums[field] += v

    # Totaux au niveau de la facture : pour les montants détaillés, on prend
    # la somme exacte de toutes les lignes (les "Totaux" du fichier ne sont
    # que des sous-totaux de page). Pour le solde de la facture, on utilise
    # la valeur "Total facture (FCFA)" extraite du fichier (dédupliquée),
    # avec la somme des lignes en repli si elle est absente.
    for field in NUMERIC_FIELDS:
        if field == "solde_facture":
            setattr(facture, field, file_totals.get(field, sums[field]))
        else:
            setattr(facture, field, sums[field])

    db.commit()

    return ImportResult(
        facture_id=facture.id,
        total_lignes=reconnus + len(non_reconnus_list),
        reconnus=reconnus,
        non_reconnus=len(non_reconnus_list),
        montant_total=montant_total,
        numeros_inconnus=non_reconnus_list[:200],
    )


@router.delete("/{facture_id}", status_code=204)
def delete_facture(facture_id: int, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(FactureTelecom).filter(FactureTelecom.id == facture_id).first()
    if not obj:
        raise HTTPException(404, "Facture introuvable")
    db.delete(obj); db.commit()


@router.get("/stats/mensuel")
def stats_mensuel(annee: int, db: Session = Depends(get_db)):
    from ..models.telephonie import CategorieSimEnum
    rows = (
        db.query(
            FactureTelecom.mois,
            NumeroSIM.categorie,
            func.sum(LigneFacture.montant).label("total"),
        )
        .join(LigneFacture, FactureTelecom.id == LigneFacture.facture_id)
        .join(NumeroSIM, LigneFacture.sim_id == NumeroSIM.id)
        .filter(FactureTelecom.annee == annee, LigneFacture.sim_id.isnot(None))
        .group_by(FactureTelecom.mois, NumeroSIM.categorie)
        .order_by(FactureTelecom.mois)
        .all()
    )
    result: dict = {}
    for r in rows:
        m = str(r.mois)
        if m not in result:
            result[m] = {}
        result[m][r.categorie] = float(r.total)
    return result
