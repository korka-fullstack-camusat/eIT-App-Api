from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
from typing import Optional
from datetime import date
from pydantic import BaseModel
from ..database import get_db
from ..models.attribution import Attribution, StatutAttribution
from ..models.materiel import Materiel, StatutMateriel
from ..schemas.attribution import AttributionCreate, AttributionOut, RestitutionCreate
from ..services.pdf_service import generate_decharge_pdf, generate_attestation_pdf, generate_recuperation_pdf
from ..models.user import User
from ..services.auth_service import require_editor


class AttributionUpdate(BaseModel):
    etat_remise: Optional[str] = None
    notes:       Optional[str] = None

router = APIRouter(prefix="/api/attributions", tags=["Attributions"])


def _get_or_404(db, attribution_id):
    obj = db.query(Attribution).options(joinedload(Attribution.materiel)).filter(Attribution.id == attribution_id).first()
    if not obj:
        raise HTTPException(404, "Attribution introuvable")
    return obj


@router.get("/stats/summary")
def attribution_stats(db: Session = Depends(get_db)):
    active    = db.query(func.count(Attribution.id)).filter(Attribution.statut == StatutAttribution.ACTIVE).scalar()
    cloturee  = db.query(func.count(Attribution.id)).filter(Attribution.statut == StatutAttribution.CLOTUREE).scalar()
    employees = db.query(func.count(func.distinct(Attribution.employee_id))).filter(Attribution.statut == StatutAttribution.ACTIVE).scalar()
    services  = (
        db.query(Attribution.employee_service, func.count(Attribution.id))
        .filter(Attribution.statut == StatutAttribution.ACTIVE)
        .group_by(Attribution.employee_service)
        .order_by(func.count(Attribution.id).desc())
        .limit(6)
        .all()
    )
    return {
        "active":           active,
        "cloturee":         cloturee,
        "employees_actifs": employees,
        "par_service":      [{"service": r[0] or "—", "count": r[1]} for r in services],
    }


@router.get("/", response_model=list[AttributionOut])
def list_attributions(
    employee_id: Optional[int] = None,
    statut: Optional[StatutAttribution] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Attribution).options(joinedload(Attribution.materiel))
    if employee_id:
        q = q.filter(Attribution.employee_id == employee_id)
    if statut:
        q = q.filter(Attribution.statut == statut)
    return q.order_by(Attribution.date_attribution.desc()).all()


@router.post("/", response_model=AttributionOut, status_code=201)
def create_attribution(data: AttributionCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    materiel = db.query(Materiel).filter(Materiel.id == data.materiel_id).first()
    if not materiel:
        raise HTTPException(404, "Matériel introuvable")
    if materiel.statut != StatutMateriel.DISPONIBLE:
        msgs = {
            StatutMateriel.ATTRIBUE:    "Ce matériel est déjà attribué",
            StatutMateriel.EN_PANNE:    "Ce matériel est en panne et ne peut pas être attribué",
            StatutMateriel.MAINTENANCE: "Ce matériel est en maintenance",
            StatutMateriel.REFORME:     "Ce matériel est réformé",
        }
        raise HTTPException(400, msgs.get(materiel.statut, "Ce matériel n'est pas disponible"))

    obj = Attribution(**data.model_dump())
    materiel.statut = StatutMateriel.ATTRIBUE
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


# ── Routes statiques déclarées AVANT /{attribution_id} ───────────────────────

class BulkAssignPayload(BaseModel):
    materiel_ids:       list[int]
    employee_id:        int
    employee_nom:       str
    employee_prenom:    Optional[str] = None
    employee_matricule: Optional[str] = None
    employee_service:   Optional[str] = None
    employee_poste:     Optional[str] = None
    date_attribution:   date
    etat_remise:        Optional[str] = None
    notes:              Optional[str] = None


@router.post("/bulk", status_code=201)
def create_bulk_attribution(
    data: BulkAssignPayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    """Assigne plusieurs matériels à un même employé en une seule opération."""
    import io
    created = []
    errors  = []
    for mid in data.materiel_ids:
        materiel = db.query(Materiel).filter(Materiel.id == mid).first()
        if not materiel:
            errors.append({"materiel_id": mid, "message": "Matériel introuvable"}); continue
        if materiel.statut != StatutMateriel.DISPONIBLE:
            errors.append({"materiel_id": mid, "message": f"Statut {materiel.statut.value} — non disponible"}); continue
        obj = Attribution(
            materiel_id        = mid,
            employee_id        = data.employee_id,
            employee_nom       = data.employee_nom,
            employee_prenom    = data.employee_prenom,
            employee_matricule = data.employee_matricule,
            employee_service   = data.employee_service,
            employee_poste     = data.employee_poste,
            date_attribution   = data.date_attribution,
            etat_remise        = data.etat_remise,
            notes              = data.notes,
        )
        materiel.statut = StatutMateriel.ATTRIBUE
        db.add(obj)
        created.append(obj)

    db.commit()
    for obj in created:
        db.refresh(obj)

    if not created:
        return {"created": 0, "errors": errors}

    attrs_with_mat = (
        db.query(Attribution)
        .options(joinedload(Attribution.materiel))
        .filter(Attribution.id.in_([o.id for o in created]))
        .all()
    )

    from ..services.template_service import template_exists, generate_attestation_from_template
    nom = f"{data.employee_prenom or ''} {data.employee_nom}".strip().replace(" ", "_")
    if template_exists("attestation"):
        doc_bytes = generate_attestation_from_template(attrs_with_mat)
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        filename   = f"attestation_{nom}.docx"
    else:
        doc_bytes  = generate_attestation_pdf(attrs_with_mat)
        media_type = "application/pdf"
        filename   = f"attestation_{nom}.pdf"

    return StreamingResponse(
        io.BytesIO(doc_bytes),
        media_type=media_type,
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "X-Created": str(len(created)),
            "X-Errors":  str(len(errors)),
        },
    )


class BulkRecuperationPayload(BaseModel):
    attribution_ids:   list[int]
    date_restitution:  date
    motif_restitution: str = "CHANGEMENT"
    notes_restitution: Optional[str] = None


@router.post("/bulk-recuperation")
def bulk_recuperation(
    data: BulkRecuperationPayload,
    db: Session = Depends(get_db),
    _: User = Depends(require_editor),
):
    """Clôture plusieurs attributions (même employé) et génère le PDF de récupération."""
    import io
    attrs = (
        db.query(Attribution)
        .options(joinedload(Attribution.materiel))
        .filter(Attribution.id.in_(data.attribution_ids))
        .all()
    )
    if not attrs:
        raise HTTPException(404, "Aucune attribution trouvée")

    for obj in attrs:
        if obj.statut == StatutAttribution.CLOTUREE:
            continue
        obj.date_restitution  = data.date_restitution
        obj.motif_restitution = data.motif_restitution
        obj.notes_restitution = data.notes_restitution
        obj.statut            = StatutAttribution.CLOTUREE
        if obj.materiel:
            obj.materiel.statut = StatutMateriel.DISPONIBLE

    db.commit()

    active = [o for o in attrs if o.materiel]
    if not active:
        raise HTTPException(400, "Aucun matériel trouvé dans les attributions")

    from ..services.template_service import template_exists, generate_recuperation_from_template
    nom = f"{active[0].employee_prenom or ''} {active[0].employee_nom}".strip().replace(" ", "_")
    if template_exists("recuperation"):
        doc_bytes  = generate_recuperation_from_template(active, date_recuperation=data.date_restitution)
        media_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        filename   = f"recuperation_{nom}.docx"
    else:
        doc_bytes  = generate_recuperation_pdf(active, date_recuperation=data.date_restitution)
        media_type = "application/pdf"
        filename   = f"recuperation_{nom}.pdf"

    return StreamingResponse(
        io.BytesIO(doc_bytes),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export-excel")
def export_attributions_excel_proxy(
    date_debut: Optional[date]              = Query(None),
    date_fin:   Optional[date]              = Query(None),
    statut:     Optional[StatutAttribution] = Query(None),
    service:    Optional[str]               = Query(None),
    search:     Optional[str]               = Query(None),
    cols:       Optional[str]               = Query(None),
    db: Session = Depends(get_db),
):
    """Proxy — défini avant /{attribution_id} pour éviter la capture par FastAPI."""
    return _generate_attributions_excel(
        date_debut=date_debut, date_fin=date_fin, statut=statut,
        service=service, search=search, cols=cols, db=db
    )


def _generate_attributions_excel(
    date_debut, date_fin, statut, cols, db, service=None, search=None
):
    """Génère un fichier Excel stylisé des attributions."""
    import io
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q = db.query(Attribution).options(joinedload(Attribution.materiel))
    if statut:     q = q.filter(Attribution.statut == statut)
    if date_debut: q = q.filter(Attribution.date_attribution >= date_debut)
    if date_fin:   q = q.filter(Attribution.date_attribution <= date_fin)
    if service:    q = q.filter(Attribution.employee_service == service)
    if search:
        like = f"%{search}%"
        q = q.outerjoin(Materiel, Attribution.materiel_id == Materiel.id).filter(
            Attribution.employee_nom.ilike(like)
            | Attribution.employee_prenom.ilike(like)
            | Attribution.employee_matricule.ilike(like)
            | Attribution.employee_service.ilike(like)
            | Materiel.marque.ilike(like)
        )
    attrs = q.order_by(Attribution.date_attribution.desc()).all()

    TYPE_LABELS = {
        "ORDINATEUR_PORTABLE": "PC Portable", "ORDINATEUR_FIXE": "PC Fixe",
        "ECRAN": "Écran", "SOURIS": "Souris", "CLAVIER": "Clavier",
        "TELEPHONE": "Téléphone", "IMPRIMANTE": "Imprimante",
        "SWITCH": "Switch", "ROUTEUR": "Routeur", "ONDULEUR": "Onduleur", "AUTRE": "Autre",
        "TABLETTE": "Tablette", "AP": "AP", "SERVEUR": "Serveur", "PARE_FEU": "Pare-feu",
    }
    STATUT_LABELS = {"ACTIVE": "Active", "CLOTUREE": "Clôturée"}
    MOTIF_LABELS  = {
        "DEPART": "Départ", "CHANGEMENT": "Changement", "PANNE": "Panne",
        "FIN_CONTRAT": "Fin contrat", "AUTRE": "Autre",
    }

    ALL_COLS = [
        ("id",          "ID",                  lambda a: a.id),
        ("employe",     "Employé",             lambda a: f"{a.employee_prenom or ''} {a.employee_nom}".strip()),
        ("matricule",   "Matricule",           lambda a: a.employee_matricule or ""),
        ("service",     "Service",             lambda a: a.employee_service or ""),
        ("poste",       "Poste",               lambda a: a.employee_poste or ""),
        ("materiel",    "Matériel",            lambda a: f"{a.materiel.marque} {a.materiel.modele or ''}".strip() if a.materiel else ""),
        ("type",        "Type",                lambda a: TYPE_LABELS.get(a.materiel.type_materiel, "") if a.materiel else ""),
        ("serie",       "N° Série",            lambda a: a.materiel.numero_serie or "" if a.materiel else ""),
        ("mac",         "Adresse MAC",         lambda a: a.materiel.adresse_mac or "" if a.materiel else ""),
        ("date_attr",   "Date Attribution",    lambda a: a.date_attribution.strftime("%d/%m/%Y") if a.date_attribution else ""),
        ("date_rest",   "Date Restitution",    lambda a: a.date_restitution.strftime("%d/%m/%Y") if a.date_restitution else ""),
        ("statut",      "Statut",              lambda a: STATUT_LABELS.get(a.statut, a.statut or "")),
        ("motif",       "Motif",               lambda a: MOTIF_LABELS.get(a.motif_restitution or "", a.motif_restitution or "")),
        ("etat_remise", "État Remise",         lambda a: a.etat_remise or ""),
        ("notes",       "Notes",               lambda a: a.notes or ""),
    ]

    selected_keys = set(cols.split(",")) if cols else {c[0] for c in ALL_COLS}
    columns = [(k, lbl, fn) for k, lbl, fn in ALL_COLS if k in selected_keys]

    # ── Styles ────────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Attributions"

    BLUE_HDR   = "1B3D6F"
    WHITE      = "FFFFFF"
    ROW_EVEN   = "EEF4FF"
    BORDER_COL = "C5D3E8"
    thin       = Side(style="thin", color=BORDER_COL)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    tc = ws.cell(row=1, column=1, value="REGISTRE DES ATTRIBUTIONS")
    tc.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    tc.fill      = PatternFill("solid", fgColor=BLUE_HDR)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Sous-titre
    parts = [f"Exporté le {datetime.now().strftime('%d/%m/%Y')}"]
    if date_debut: parts.append(f"Du {date_debut.strftime('%d/%m/%Y')}")
    if date_fin:   parts.append(f"au {date_fin.strftime('%d/%m/%Y')}")
    if statut:     parts.append(f"Statut : {STATUT_LABELS.get(statut, statut)}")
    parts.append(f"{len(attrs)} attribution(s)")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))
    sc = ws.cell(row=2, column=1, value="  |  ".join(parts))
    sc.font      = Font(name="Calibri", italic=True, size=9, color="5B7DB1")
    sc.fill      = PatternFill("solid", fgColor="D9E6F7")
    sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 6

    # En-têtes
    HDR_ROW = 4
    for ci, (_, lbl, _fn) in enumerate(columns, start=1):
        c = ws.cell(row=HDR_ROW, column=ci, value=lbl)
        c.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        c.fill      = PatternFill("solid", fgColor=BLUE_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border_all
    ws.row_dimensions[HDR_ROW].height = 24
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)

    # Données
    for ri, a in enumerate(attrs):
        row_num = HDR_ROW + 1 + ri
        row_fill = PatternFill("solid", fgColor=(ROW_EVEN if ri % 2 == 0 else WHITE))
        for ci, (_, _lbl, fn) in enumerate(columns, start=1):
            c = ws.cell(row=row_num, column=ci, value=fn(a))
            c.fill      = row_fill
            c.font      = Font(name="Calibri", size=10)
            c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
            c.border    = border_all
        ws.row_dimensions[row_num].height = 18

    # Largeurs auto
    for ci, (_, lbl, fn) in enumerate(columns, start=1):
        max_len = max((len(str(fn(a))) for a in attrs), default=0)
        ws.column_dimensions[get_column_letter(ci)].width = max(len(lbl) + 2, min(max_len + 2, 45))

    ws.oddFooter.center.text = "&\"Calibri\"&8 CAMUSAT — Attributions  |  Page &P / &N"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = "attributions"
    if date_debut: fname += f"_du_{date_debut}"
    if date_fin:   fname += f"_au_{date_fin}"
    fname += ".xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/{attribution_id}", response_model=AttributionOut)
def get_attribution(attribution_id: int, db: Session = Depends(get_db)):
    return _get_or_404(db, attribution_id)


@router.patch("/{attribution_id}", response_model=AttributionOut)
def update_attribution(attribution_id: int, data: AttributionUpdate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = db.query(Attribution).filter(Attribution.id == attribution_id).first()
    if not obj:
        raise HTTPException(404, "Attribution introuvable")
    if data.etat_remise is not None:
        obj.etat_remise = data.etat_remise
    if data.notes is not None:
        obj.notes = data.notes
    db.commit()
    db.refresh(obj)
    # Recharger avec le matériel pour la réponse
    return db.query(Attribution).options(joinedload(Attribution.materiel)).filter(Attribution.id == attribution_id).first()


@router.post("/{attribution_id}/restitution", response_model=AttributionOut)
def restituer(attribution_id: int, data: RestitutionCreate, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    obj = _get_or_404(db, attribution_id)
    if obj.statut == StatutAttribution.CLOTUREE:
        raise HTTPException(400, "Attribution déjà clôturée")

    obj.date_restitution  = data.date_restitution
    obj.motif_restitution = data.motif_restitution
    obj.notes_restitution = data.notes_restitution
    obj.statut            = StatutAttribution.CLOTUREE
    obj.materiel.statut   = StatutMateriel.DISPONIBLE
    db.commit()
    db.refresh(obj)
    return obj


@router.get("/{attribution_id}/decharge")
def download_decharge(attribution_id: int, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import io
    obj = _get_or_404(db, attribution_id)
    pdf_bytes = generate_decharge_pdf(obj)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=decharge_{attribution_id}.pdf"},
    )


@router.get("/attestation/employee/{employee_id}")
def attestation_employee(employee_id: int, db: Session = Depends(get_db)):
    from fastapi.responses import StreamingResponse
    import io
    attrs = (
        db.query(Attribution)
        .options(joinedload(Attribution.materiel))
        .filter(Attribution.employee_id == employee_id, Attribution.statut == StatutAttribution.ACTIVE)
        .order_by(Attribution.date_attribution.asc())
        .all()
    )
    if not attrs:
        raise HTTPException(404, "Aucune attribution active pour cet employé")
    pdf_bytes = generate_attestation_pdf(attrs)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename=attestation_{employee_id}.pdf"},
    )


@router.get("/employee/{employee_id}", response_model=list[AttributionOut])
def attributions_par_employee(employee_id: int, db: Session = Depends(get_db)):
    return (
        db.query(Attribution)
        .options(joinedload(Attribution.materiel))
        .filter(Attribution.employee_id == employee_id)
        .order_by(Attribution.date_attribution.desc())
        .all()
    )


@router.post("/{attribution_id}/transferer", status_code=201)
def transferer(attribution_id: int, data: dict, db: Session = Depends(get_db), _: User = Depends(require_editor)):
    """Clôture l'attribution en cours et crée immédiatement la suivante."""
    from pydantic import BaseModel
    from typing import Optional as Opt
    from datetime import date as dt_date

    old = _get_or_404(db, attribution_id)
    if old.statut == StatutAttribution.CLOTUREE:
        raise HTTPException(400, "Attribution déjà clôturée")

    # Clôture
    old.date_restitution  = dt_date.fromisoformat(data["date_restitution"])
    old.motif_restitution = data.get("motif_restitution", "CHANGEMENT")
    old.notes_restitution = data.get("notes_restitution")
    old.statut            = StatutAttribution.CLOTUREE

    # Nouvelle attribution
    new_attr = Attribution(
        materiel_id        = old.materiel_id,
        employee_id        = data["employee_id"],
        employee_nom       = data["employee_nom"],
        employee_prenom    = data.get("employee_prenom"),
        employee_matricule = data.get("employee_matricule"),
        employee_service   = data.get("employee_service"),
        employee_poste     = data.get("employee_poste"),
        date_attribution   = dt_date.fromisoformat(data["date_attribution"]),
        etat_remise        = data.get("etat_remise"),
        notes              = data.get("notes"),
        statut             = StatutAttribution.ACTIVE,
    )
    db.add(new_attr)
    db.commit()
    db.refresh(new_attr)
    return {"old_id": attribution_id, "new_id": new_attr.id}


@router.get("/export")
def export_attributions_csv(
    date_debut: Optional[date] = Query(None, description="Date début (YYYY-MM-DD)"),
    date_fin:   Optional[date] = Query(None, description="Date fin (YYYY-MM-DD)"),
    statut:     Optional[StatutAttribution] = Query(None),
    db: Session = Depends(get_db),
):
    """Export CSV des attributions avec filtre optionnel par période."""
    import csv, io
    from fastapi.responses import StreamingResponse

    q = db.query(Attribution).options(joinedload(Attribution.materiel))
    if statut:
        q = q.filter(Attribution.statut == statut)
    if date_debut:
        q = q.filter(Attribution.date_attribution >= date_debut)
    if date_fin:
        q = q.filter(Attribution.date_attribution <= date_fin)
    attrs = q.order_by(Attribution.date_attribution.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow([
        "ID", "Employé", "Matricule", "Service", "Poste",
        "Matériel", "Type", "Marque", "Modèle", "N° Série", "Adresse MAC",
        "Date Attribution", "Date Restitution", "Statut", "Motif Restitution",
        "État Remise", "Notes",
    ])
    for a in attrs:
        m = a.materiel
        writer.writerow([
            a.id,
            f"{a.employee_prenom or ''} {a.employee_nom}".strip(),
            a.employee_matricule or "",
            a.employee_service or "",
            a.employee_poste or "",
            f"{m.marque} {m.modele or ''}".strip() if m else "",
            m.type_materiel if m else "",
            m.marque if m else "",
            m.modele or "" if m else "",
            m.numero_serie or "" if m else "",
            m.adresse_mac or "" if m else "",
            a.date_attribution.isoformat() if a.date_attribution else "",
            a.date_restitution.isoformat() if a.date_restitution else "",
            a.statut,
            a.motif_restitution or "",
            a.etat_remise or "",
            a.notes or "",
        ])

    content = "﻿" + output.getvalue()  # BOM UTF-8 pour Excel
    fname = "attributions"
    if date_debut:
        fname += f"_{date_debut}"
    if date_fin:
        fname += f"_au_{date_fin}"
    fname += ".csv"

    return StreamingResponse(
        iter([content.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.post("/export-template")
async def export_with_template(
    file: UploadFile = File(...),
    date_debut: Optional[date] = Query(None),
    date_fin:   Optional[date] = Query(None),
    statut:     Optional[StatutAttribution] = Query(None),
    db: Session = Depends(get_db),
):
    """Remplit le modèle CSV uploadé avec les données des attributions."""
    import csv, io
    from fastapi.responses import StreamingResponse

    # Lire et parser le modèle uploadé pour récupérer les en-têtes
    content = await file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.reader(io.StringIO(text), delimiter=";")
    try:
        headers = [h.strip() for h in next(reader)]
    except StopIteration:
        raise HTTPException(400, "Le fichier modèle est vide")

    # Mapping en-tête → champ d'attribution
    FIELD_MAP: dict[str, str] = {
        "id": "id",
        "employe": "employee_full",
        "employé": "employee_full",
        "nom": "employee_nom",
        "prenom": "employee_prenom",
        "prénom": "employee_prenom",
        "matricule": "employee_matricule",
        "service": "employee_service",
        "poste": "employee_poste",
        "materiel": "materiel_label",
        "matériel": "materiel_label",
        "type": "materiel_type",
        "marque": "materiel_marque",
        "modele": "materiel_modele",
        "modèle": "materiel_modele",
        "n° série": "materiel_serie",
        "n serie": "materiel_serie",
        "numero serie": "materiel_serie",
        "numéro série": "materiel_serie",
        "adresse mac": "materiel_mac",
        "adresse ip": "materiel_mac",
        "date attribution": "date_attribution",
        "date d'attribution": "date_attribution",
        "date restitution": "date_restitution",
        "statut": "statut",
        "motif": "motif_restitution",
        "motif restitution": "motif_restitution",
        "etat remise": "etat_remise",
        "état remise": "etat_remise",
        "notes": "notes",
    }

    # Charger les attributions
    q = db.query(Attribution).options(joinedload(Attribution.materiel))
    if statut:
        q = q.filter(Attribution.statut == statut)
    if date_debut:
        q = q.filter(Attribution.date_attribution >= date_debut)
    if date_fin:
        q = q.filter(Attribution.date_attribution <= date_fin)
    attrs = q.order_by(Attribution.date_attribution.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(headers)

    for a in attrs:
        m = a.materiel
        data: dict[str, str] = {
            "id":               str(a.id),
            "employee_full":    f"{a.employee_prenom or ''} {a.employee_nom}".strip(),
            "employee_nom":     a.employee_nom or "",
            "employee_prenom":  a.employee_prenom or "",
            "employee_matricule": a.employee_matricule or "",
            "employee_service": a.employee_service or "",
            "employee_poste":   a.employee_poste or "",
            "materiel_label":   f"{m.marque} {m.modele or ''}".strip() if m else "",
            "materiel_type":    m.type_materiel if m else "",
            "materiel_marque":  m.marque if m else "",
            "materiel_modele":  m.modele or "" if m else "",
            "materiel_serie":   m.numero_serie or "" if m else "",
            "materiel_mac":     m.adresse_mac or "" if m else "",
            "date_attribution": a.date_attribution.isoformat() if a.date_attribution else "",
            "date_restitution": a.date_restitution.isoformat() if a.date_restitution else "",
            "statut":           a.statut or "",
            "motif_restitution": a.motif_restitution or "",
            "etat_remise":      a.etat_remise or "",
            "notes":            a.notes or "",
        }
        row = []
        for h in headers:
            key = FIELD_MAP.get(h.lower().strip(), "")
            row.append(data.get(key, ""))
        writer.writerow(row)

    result = "﻿" + output.getvalue()
    return StreamingResponse(
        iter([result.encode("utf-8")]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": "attachment; filename=attributions_filled.csv"},
    )


@router.get("/materiel/{materiel_id}", response_model=list[AttributionOut])
def attributions_par_materiel(materiel_id: int, db: Session = Depends(get_db)):
    """Historique complet (actives + clôturées) d'un matériel."""
    return (
        db.query(Attribution)
        .options(joinedload(Attribution.materiel))
        .filter(Attribution.materiel_id == materiel_id)
        .order_by(Attribution.date_attribution.desc())
        .all()
    )
